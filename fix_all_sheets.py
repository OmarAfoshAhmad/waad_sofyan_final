import os
import glob
from openpyxl import load_workbook

base_dir = r'D:\tba_waad_system-main_success\tba_waad_system-main'
folders = [
    'Unified_Benefit_Tables',
    'Unified_Benefit_Tables_Ready',
    'Unified_Benefit_Tables_Ready_V2',
    'Unified_Benefit_Tables_Ready_V3'
]

total_fixed = 0

for folder in folders:
    folder_path = os.path.join(base_dir, folder)
    if not os.path.exists(folder_path):
        continue
    
    excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    for file in excel_files:
        try:
            wb = load_workbook(file)
            changed = False
            for sheet in wb.sheetnames:
                if sheet == 'قواعد التغطية':
                    wb[sheet].title = 'قواعد_التغطية'
                    changed = True
            
            if changed:
                wb.save(file)
                print(f"Fixed sheet name in: {os.path.basename(folder)}/{os.path.basename(file)}")
                total_fixed += 1
        except Exception as e:
            print(f"Error processing {os.path.basename(file)}: {e}")

if total_fixed == 0:
    print("All files are already correct.")
else:
    print(f"Total files fixed: {total_fixed}")
