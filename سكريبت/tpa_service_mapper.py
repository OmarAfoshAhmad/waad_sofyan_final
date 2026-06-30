#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  TPA Service Mapper  |  مُرتِّب ومُطابِق قوائم الأسعار لنظام التأمين الطبي
=============================================================================
الوظيفة:
  يأخذ ملف Excel "مرجعي" (كتالوج الخدمات الرسمي) + أي قائمة أسعار "فوضوية"
  مُرسَلة من مستشفى/صيدلية/مختبر، ثم يُنتج ملف Excel مُرتَّب وجاهز للاستيراد
  داخل نظام TPA، بنفس أسماء أعمدة المرجع، مع تعبئة:
      - service_code      (الكود)
      - main_category     (التصنيف الرئيسي: إيواء / عيادات خارجية)
      - sub_category      (التصنيف الفرعي: CAT0xx)
      - confidence_score  (درجة الثقة في المطابقة)

  الخدمات التي لا يجد لها تطابقًا كافيًا تُوضَع في ورقة منفصلة "Need Review"
  مع أفضل اقتراح، ليراجعها موظف بشري.

طرق المطابقة (بالترتيب):
  1) Exact match            - تطابق تام بعد التنظيف
  2) Synonyms               - عبر قاموس المرادفات الطبية (قابل للتوسعة)
  3) Arabic/English Normalize - توحيد الكتابة العربية والإنجليزية
  4) Fuzzy Matching         - تطابق تقريبي (rapidfuzz إن وُجد، وإلا difflib)

طريقة التشغيل:
  python tpa_service_mapper.py \
      --reference  Price_List_Contract_1_Output.xlsx \
      --input      new_hospital_list.xlsx \
      --output     result.xlsx \
      --synonyms   medical_synonyms.json \
      --threshold  85

  (المعاملات --synonyms و --threshold اختيارية)
=============================================================================
"""

import argparse
import json
import os
import re
import sys
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import ingest  # محرك الابتلاع متعدد الصيغ (Excel/PDF/PowerPoint/CSV)

# ---------------------------------------------------------------------------
# محرك المطابقة التقريبية: يفضّل rapidfuzz، ويسقط تلقائيًا إلى difflib
# ---------------------------------------------------------------------------
try:
    from rapidfuzz import fuzz
    from rapidfuzz import process as rf_process    # الأفضل أداءً ودقة
    _FUZZ_ENGINE = "rapidfuzz"
except ImportError:                                # خطة بديلة بدون أي تثبيت
    _FUZZ_ENGINE = "index"


# ===========================================================================
# 1) إعدادات افتراضية  (Defaults)
# ===========================================================================
DEFAULT_THRESHOLD = 85          # أقل درجة ثقة لقبول المطابقة تلقائيًا (0-100)
EXACT_SCORE = 100               # درجة التطابق التام

# الأعمدة المعتمدة في المخرجات (نفس أسماء المرجع تمامًا)
COL_NAME  = "service_name / اسم الخدمة ★"
COL_CODE  = "service_code / الكود"
COL_PRICE = "contract_price / سعر العقد"
COL_MAIN  = "main_category / التصنيف الرئيسي"
COL_SUB   = "sub_category / البند (التصنيف الفرعي)"
COL_NOTE  = "notes / ملاحظات"
STD_COLS  = [COL_NAME, COL_CODE, COL_PRICE, COL_MAIN, COL_SUB, COL_NOTE]

# أعمدة مساعدة للمراجع (تأتي بعد أعمدة الاعتماد الستة)
COL_REASON  = "سبب النتيجة / Reason"
COL_APPROVE = "اعتماد المراجع (نعم/لا)"



# ===========================================================================
# 2) التنظيف والتوحيد  (Normalization)
# ===========================================================================
_AR_DIACRITICS = re.compile(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]")
_TATWEEL = re.compile(r"\u0640")
_NON_KEEP = re.compile(r"[^0-9a-z\u0621-\u064A\s]")   # احتفظ: أرقام/لاتيني/عربي/مسافة
_MULTISPACE = re.compile(r"\s+")


def clean_cell(v) -> str:
    """يحوّل أي قيمة خلية إلى نص نظيف، ويعامل الفراغ/nan/None كنص فارغ."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


# كلمات وصف عامة تُزال قبل المطابقة (لا تحمل معنى تمييزيًا للخدمة)
_STOPWORDS = {"serum", "plasma", "plain", "playn", "sample", "sampel",
              "specimen", "test", "in", "by", "for", "the", "of",
              "عينه", "عينة", "في", "تحليل"}


def normalize(text) -> str:
    """تنظيف وتوحيد النص (عربي + إنجليزي) لجعله قابلًا للمقارنة."""
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKC", s)      # توحيد الأشكال (أرقام/حروف)
    s = s.lower()                             # توحيد اللاتيني
    s = _AR_DIACRITICS.sub("", s)             # إزالة التشكيل
    s = _TATWEEL.sub("", s)                   # إزالة التطويل ـــ
    # توحيد الألف والهمزات والياء والتاء المربوطة
    s = re.sub(r"[إأآٱ]", "ا", s)
    s = s.replace("ى", "ي").replace("ئ", "ي").replace("ؤ", "و")
    s = s.replace("ة", "ه")
    s = _NON_KEEP.sub(" ", s)                 # إزالة الرموز والترقيم
    s = _MULTISPACE.sub(" ", s).strip()       # دمج المسافات
    # أزل كلمات الوصف العامة (مع الإبقاء على كلمة واحدة على الأقل)
    toks = [t for t in s.split() if t not in _STOPWORDS]
    return " ".join(toks) if toks else s


# ===========================================================================
# 3) قاموس المرادفات  (Synonyms)
# ===========================================================================
def load_synonyms(path):
    """يحمّل قاموس المرادفات ويبني خريطة (مرادف مُنظَّف -> الصيغة المعتمدة المُنظَّفة)."""
    mapping = {}
    if not path or not os.path.exists(path):
        return mapping
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    for canonical, variants in raw.items():
        if canonical.startswith("_"):         # تجاهل مفاتيح الوصف (_description …)
            continue
        canon_n = normalize(canonical)
        if not canon_n:
            continue
        mapping[canon_n] = canon_n
        for v in variants:
            vn = normalize(v)
            if vn:
                mapping[vn] = canon_n
    return mapping


def apply_synonyms(norm_text, syn_map):
    """يستبدل أي كلمة/عبارة لها مرادف معروف بصيغتها المعتمدة (مطابقة كلمات كاملة)."""
    if not syn_map or not norm_text:
        return norm_text
    words = norm_text.split()
    out = []
    i = 0
    n = len(words)
    while i < n:
        replaced = False
        # جرّب عبارات من 3 كلمات ثم 2 ثم 1 (الأطول أولًا)
        for span in (3, 2, 1):
            if i + span <= n:
                phrase = " ".join(words[i:i + span])
                if phrase in syn_map:
                    out.append(syn_map[phrase])
                    i += span
                    replaced = True
                    break
        if not replaced:
            out.append(words[i])
            i += 1
    return _MULTISPACE.sub(" ", " ".join(out)).strip()


# ===========================================================================
# 4) تحميل المرجع وبناء الفهرس  (Reference index)
# ===========================================================================
def build_reference(ref_path, syn_map):
    """يقرأ المرجع ويبني فهرسًا: لكل اسم مُنظَّف -> (code, main, sub, الاسم الأصلي)."""
    df = pd.read_excel(ref_path, dtype=str)
    # تطابق الأعمدة بالاسم؛ وإن اختلفت قليلًا نأخذ أول 6 أعمدة بالترتيب
    cols = list(df.columns)
    if not all(c in cols for c in STD_COLS):
        if len(cols) >= 5:
            rename = dict(zip(cols[:6], STD_COLS[:len(cols[:6])]))
            df = df.rename(columns=rename)
        else:
            raise ValueError("ملف المرجع لا يحتوي على الأعمدة المتوقعة.")

    exact = {}      # اسم مُنظَّف -> سجل
    exact_syn = {}  # اسم مُنظَّف + مرادفات -> سجل
    choices = {}    # اسم مُنظَّف -> الاسم المُنظَّف (لقائمة المطابقة التقريبية)

    for _, row in df.iterrows():
        name = row.get(COL_NAME)
        if name is None or str(name).strip() == "" or str(name) == "nan":
            continue
        rec = {
            "name": clean_cell(name),
            "code": clean_cell(row.get(COL_CODE)),
            "main": clean_cell(row.get(COL_MAIN)),
            "sub":  clean_cell(row.get(COL_SUB)),
        }
        nn = normalize(name)
        if not nn:
            continue
        # أول ظهور يفوز (المرجع قد يكرر نفس الاسم بأكواد مختلفة)
        exact.setdefault(nn, rec)
        choices.setdefault(nn, nn)
        nsyn = apply_synonyms(nn, syn_map)
        exact_syn.setdefault(nsyn, rec)
        choices.setdefault(nsyn, nsyn)

    choice_list = list(choices.keys())
    return df, exact, exact_syn, choices, choice_list


# ===========================================================================
# 4.5) المُصنِّف الطبي  (Medical category classifier)
# ===========================================================================
# قواعد بالأولوية: لكل كود تصنيف كلماتٌ مفتاحية طبية (الأكثر تحديدًا أولًا).
# الكلمات تُكتب بصيغة مبسّطة وتُطبَّع آليًا قبل المطابقة.
_CATEGORY_RULES = [
    ("CAT024", ["مقطعي", "مقطعيه", "رنين", "mri", "ct scan", "ct brain",
                "computed tomography", "magnetic resonance", "ct", "hrct"]),
    ("CAT031", ["تقويم", "زراعه سن", "زراعه عظم", "تلبيس", "طاقم", "تاج",
                "جسر", "denture", "crown", "implant", "orthodont", "veneer",
                "bridge", "prosth", "زركون", "خزف", "وتد", "قشره"]),
    ("CAT028", ["اسنان", "لثه", "حشو", "خلع", "عصب سن", "tooth", "teeth",
                "dental", "extraction", "filling", "gingiv", "root canal",
                "scaling", "caries", "ضرس", "تلميع", "فلورايد"]),
    ("CAT029", ["عيون", "نظر", "شبكيه", "قرنيه", "جفن", "حدقه", "eye",
                "retina", "ocular", "cornea", "optic", "cataract", "glaucoma",
                "fundus", "قاع العين", "عدسه", "بصر", "ophthalm"]),
    ("CAT027", ["علاج طبيعي", "تاهيل", "physiother", "rehabilit", "تدليك",
                "massage", "موجات تصادميه", "شد العنق", "شد الظهر"]),
    ("CAT021", ["قيصري", "قيصريه", "ولاده", "cesarean", "caesarean",
                "delivery", "مشيمه", "placenta", "مخاض", "nvd", "حمل"]),
    ("CAT002", ["عنايه فايقه", "عنايه مركزه", "عنايه القلب", "icu", "ccu",
                "intensive care", "عنايه حديثي", "حضانه", "nicu"]),
    ("CAT004", ["تخدير", "تخذير", "بنج", "anesth", "anaesth", "sedation"]),
    ("CAT007", ["اسعاف", "ambulance"]),
    ("CAT026", ["جهاز", "منظار", "قسطره", "قسطار", "انبوب", "scope",
                "endoscop", "catheter", "ventilator", "device", "cannula",
                "دعامه", "tube", "drain", "غسيل كلوي", "dialysis"]),
    ("CAT025", ["حقنه", "حقن", "ادويه", "محاليل", "injection", "infusion",
                "تطعيم", "لقاح", "vaccine", "مغذي", "سوائل", "تسريب"]),
    ("CAT003", ["عمليه", "جراحه", "استئصال", "بتر", "resection", "ectomy",
                "otomy", "operation", "surgery", "surgical", "excision",
                "biopsy", "خزعه", "تفتيت", "ترقيع", "زرع"]),
    ("CAT023", ["تحليل", "فحص", "مزرعه", "حساسيه", "cbc", "esr", "culture",
                "serum", "blood", "urine", "stool", "semen", "hormone",
                "هرمون", "antibody", "اجسام مضاده", "antigen", "igg", "igm",
                "pcr", "dna", "vitamin", "فيتامين", "سكر", "glucose",
                "cholesterol", "دهون", "كرياتينين", "creatinine", "حمض",
                "level", "count", "screen", "titer", "ferritin", "psa",
                "cea", "tsh", "ايكو", "echo", "دوبلر", "doppler", "موجات",
                "ultrasound", "سونار", "uss", "اشعه", "x ray", "كشف",
                "استشاري", "اخصاءي", "مراجعه", "consultation", "ecg",
                "تخطيط", "صوره"]),
]


class CategoryClassifier:
    """يتعلّم أكواد التصنيف الفعلية من المرجع، ويصنّف أي خدمة بقواعد طبية."""

    def __init__(self, df_ref):
        self.sub_by_key = {}     # CAT023 -> النص الكامل للتصنيف الفرعي
        self.main_by_key = {}    # CAT023 -> التصنيف الرئيسي
        subs = df_ref[COL_SUB].dropna().astype(str)
        mains = df_ref[COL_MAIN].astype(str)
        for sub, main in zip(subs, mains):
            m = re.match(r"\s*(CAT\d+)", sub)
            if m:
                key = m.group(1)
                self.sub_by_key.setdefault(key, sub.strip())
                self.main_by_key.setdefault(key, clean_cell(main))
        # التصنيف الافتراضي = الأكثر شيوعًا في المرجع
        counts = df_ref[COL_SUB].dropna().astype(str).map(
            lambda s: (re.match(r"\s*(CAT\d+)", s) or [None, None])[1]
            if re.match(r"\s*(CAT\d+)", s) else None)
        self.default_key = counts.value_counts().idxmax() if len(counts) else None
        # طبِّع الكلمات المفتاحية مرة واحدة
        self.rules = [(k, [normalize(w) for w in kws if normalize(w)])
                      for k, kws in _CATEGORY_RULES if k in self.sub_by_key]

    @staticmethod
    def _stem(tok):
        """يجرّد أدوات التعريف من كلمة عربية (الـ/وال/بال...) لتسهيل المطابقة."""
        for pre in ("وال", "فال", "بال", "كال", "لل", "ال"):
            if tok.startswith(pre) and len(tok) > len(pre):
                return tok[len(pre):]
        if len(tok) > 1 and tok[0] in "وفبكل":
            return tok[1:]
        return tok

    def _ar_token_match(self, kw, stems):
        return kw in stems

    def _hit(self, kw, text, tokens, stems):
        if " " in kw:                       # عبارة
            if kw.isascii():
                return kw in text           # لاتينية: مطابقة جزئية
            # عربية: كل كلماتها موجودة كجذور (يتجاوز «ال» الداخلية)
            return all(self._stem(w) in stems for w in kw.split())
        if kw.isascii():
            if len(kw) >= 5:                # لاحقة/مقطع (ectomy, scope) → جزئي
                return kw in text
            return kw in tokens             # كلمة قصيرة (ct, mri) → كاملة
        return kw in stems                  # عربي مفرد → جذر كامل

    def classify(self, raw_text):
        """يعيد (main, sub, catkey) — تصنيف تقديري بالخبرة الطبية."""
        t = normalize(raw_text)
        if not t:
            return "", "", None
        toks = set(t.split())
        stems = {self._stem(x) for x in toks}
        # 1) قواعد التصنيف الطبية (الأكثر تحديدًا أولًا)
        for key, kws in self.rules:
            for kw in kws:
                if self._hit(kw, t, toks, stems):
                    return self.main_by_key.get(key, ""), self.sub_by_key[key], key
        # 2) الإيواء/الإقامة (غرف، أسرّة) → التصنيف الرئيسي «إيواء»
        if {"اقامه", "غرفه", "سرير", "مبيت", "ايواء", "جناح"} & stems:
            return "إيواء", "", None
        # 3) الافتراضي = الأكثر شيوعًا بالمرجع
        if self.default_key:
            return (self.main_by_key.get(self.default_key, ""),
                    self.sub_by_key[self.default_key], self.default_key)
        return "", "", None


# ===========================================================================
# 5) المطابقة التقريبية  (Fuzzy)
# ===========================================================================
import difflib
from collections import defaultdict


class FuzzyIndex:
    """فهرس كلمات مقلوب لمطابقة تقريبية سريعة بدون أي مكتبات خارجية.

    يحجب المقارنات إلى المرشّحين الذين يشاركون كلمة واحدة على الأقل، ثم
    يرتّبهم بمعامل تشابه المجموعات (Jaccard) ويصقل الأفضل بـ difflib.
    """

    def __init__(self, choice_list):
        self.choices = choice_list
        self.tokens = [set(c.split()) for c in choice_list]
        self.inverted = defaultdict(list)
        for i, toks in enumerate(self.tokens):
            for t in toks:
                self.inverted[t].append(i)

    def best(self, query_norm, topk=5):
        if not query_norm:
            return None, 0, {}
        q = set(query_norm.split())
        if not q:
            return None, 0, {}
        # اجمع المرشّحين الذين يشاركون كلمة على الأقل
        cand = set()
        for t in q:
            cand.update(self.inverted.get(t, ()))
        if not cand:
            return None, 0, {}
        # رتّب بـ Jaccard ثم اصقل أفضل القلّة بـ difflib
        scored = []
        for i in cand:
            inter = len(q & self.tokens[i])
            union = len(q | self.tokens[i])
            scored.append((inter / union if union else 0, i))
        scored.sort(reverse=True)
        best_score, best_idx, best_comp = 0, None, {}
        for jac, i in scored[:topk]:
            seq = difflib.SequenceMatcher(None, query_norm, self.choices[i]).ratio()
            inter = len(q & self.tokens[i])
            # الاحتواء الجزئي — يُطبَّق فقط عندما يكون لكلا الطرفين كلمتان فأكثر
            # (يمنع مطابقة جملة طويلة بمدخل مرجعي مفرد عام مثل "IgM")
            if len(q) >= 2 and len(self.tokens[i]) >= 2:
                contain = inter / min(len(q), len(self.tokens[i]))
            else:
                contain = 0
            sc = max(jac, seq, contain * 0.9) * 100
            if sc > best_score:
                best_score, best_idx = sc, i
                best_comp = {"jac": round(jac * 100), "seq": round(seq * 100),
                             "contain": round(contain * 100)}
        return (self.choices[best_idx] if best_idx is not None else None,
                round(best_score, 1), best_comp)


def fuzzy_best(query_norm, fidx, ref_lookup):
    """يعيد (السجل, الدرجة, الاسم_المطابق, مكوّنات الدرجة)."""
    if not query_norm:
        return None, 0, None, {}
    if _FUZZ_ENGINE == "rapidfuzz":
        best = rf_process.extractOne(query_norm, fidx.choices,
                                  scorer=fuzz.token_set_ratio)
        if best is None:
            return None, 0, None, {}
        matched_norm, score = best[0], round(best[1], 1)
        comp = {"jac": "", "seq": "", "contain": ""}
    else:
        matched_norm, score, comp = fidx.best(query_norm)
        if matched_norm is None:
            return None, 0, None, {}
    return ref_lookup.get(matched_norm), score, matched_norm, comp


def _signal_text(comp):
    """يصف أي إشارة رفعت الدرجة (كلمات/حروف/احتواء)."""
    if not comp or comp.get("jac") == "":
        return "تشابه نصّي"
    jac, seq, con = comp.get("jac", 0), comp.get("seq", 0), comp.get("contain", 0)
    top = max(jac, seq, con * 0.9)
    if con * 0.9 == top and con >= 60:
        return "كلمات الخدمة جزء من اسم بالمرجع (احتواء)"
    if jac >= seq:
        return f"تطابق في الكلمات ({jac}%)"
    return f"تشابه في الحروف ({seq}%) — قد تختلف كلمة، تحقّق أنها لا تغيّر المعنى"


def build_reason(res, threshold, is_review):
    """يبني عبارة «سبب النتيجة بنظرة» للمراجع."""
    method, score, comp = res["method"], res["score"], res.get("comp", {})
    if "تطابق تام" in method:
        return "✔ تطابق تام مع المرجع"
    if "مرادف" in method:
        return "✔ مطابقة عبر مرادف معروف"
    # تقريبي
    if is_review:
        if score == 0 or not res["matched"]:
            return "✖ لا يوجد ما يقابلها في المرجع — غالبًا غير متعاقد عليها"
        if score < 50:
            return f"بعيدة عن كل خدمات المرجع ({score}%) — راجِع: قد تكون غير موجودة"
        if score < 70:
            return (f"صياغة مختلفة/كلمات زائدة ({score}%) — قد تكون صحيحة، "
                    f"أكّد الاقتراح أو عدّله")
        return f"قريبة جدًا من الاقتراح ({score}%) — راجِع وأكّد يدويًا"
    # مقبولة في النتيجة
    sig = _signal_text(comp)
    if score >= 90:
        return f"موثوق ({score}%): {sig}"
    return f"⚠ ثقة حدّية ({score}%): {sig} — يُستحسن نظرة سريعة"


def _ig_class(text):
    """يستخرج فئة الجسم المضاد (igg/igm/iga/ige) إن وُجدت."""
    found = set(re.findall(r"\big[gma e]\b|\bigg\b|\bigm\b|\biga\b|\bige\b", text))
    cls = set()
    for f in found:
        f = f.replace(" ", "")
        if f in ("igg", "igm", "iga", "ige"):
            cls.add(f)
    return cls


def match_service(raw_name, syn_map, exact, exact_syn, fidx, ref_lookup):
    """يطابق خدمة واحدة ويعيد قاموس النتيجة."""
    nn = normalize(raw_name)
    nsyn = apply_synonyms(nn, syn_map)

    # 1) تطابق تام
    if nn in exact:
        return _result(exact[nn], EXACT_SCORE, "exact / تطابق تام",
                       exact[nn]["name"], {})
    # 2) تطابق تام بعد المرادفات
    if nsyn in exact_syn:
        return _result(exact_syn[nsyn], EXACT_SCORE, "synonym / مرادف",
                       exact_syn[nsyn]["name"], {})
    # 3+4) تقريبي (نجرّب النص العادي والنص بعد المرادفات ونأخذ الأعلى)
    rec1, sc1, m1, c1 = fuzzy_best(nn,   fidx, ref_lookup)
    rec2, sc2, m2, c2 = fuzzy_best(nsyn, fidx, ref_lookup)
    if sc2 > sc1:
        rec, sc, matched, comp = rec2, sc2, m2, c2
    else:
        rec, sc, matched, comp = rec1, sc1, m1, c1

    # حارس فئات الأجسام المضادة: لا تَقبل خلط IgG/IgM/IgA/IgE
    if rec is not None:
        q_cls = _ig_class(nn) or _ig_class(nsyn)
        r_cls = _ig_class(matched or "")
        if q_cls and r_cls and not (q_cls & r_cls):
            sc = min(sc, 60)   # اخفض الدرجة لتذهب للمراجعة البشرية

    return _result(rec, sc, "fuzzy / تقريبي",
                   rec["name"] if rec else "", comp)


def _result(rec, score, method, matched_name, comp):
    if rec is None:
        return {"code": "", "main": "", "sub": "", "score": 0,
                "method": "none / لا يوجد", "matched": "", "comp": {}}
    return {"code": rec["code"], "main": rec["main"], "sub": rec["sub"],
            "score": score, "method": method, "matched": matched_name,
            "comp": comp}


# ===========================================================================
# 7) كتابة المخرجات  (Excel output with formatting)
# ===========================================================================
from openpyxl.worksheet.datavalidation import DataValidation

HDR_FILL   = PatternFill("solid", fgColor="1F4E78")   # أزرق غامق للعناوين
IMPORT_FILL = PatternFill("solid", fgColor="1F6E43")  # أخضر = الأعمدة الست للاعتماد
HELPER_FILL = PatternFill("solid", fgColor="7F7F7F")  # رمادي = أعمدة مساعدة
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HIGH_FILL  = PatternFill("solid", fgColor="C6EFCE")   # أخضر = ثقة عالية
MID_FILL   = PatternFill("solid", fgColor="FFEB9C")   # أصفر = ثقة متوسطة
LOW_FILL   = PatternFill("solid", fgColor="FFC7CE")   # أحمر = ثقة منخفضة
BODY_FONT  = Font(name="Arial", size=10)


def _style_sheet(ws, n_cols, conf_col=None, import_cols=0):
    """import_cols: عدد الأعمدة الأولى التي تطابق ملف الاعتماد (تُلوَّن مميَّزة)."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = IMPORT_FILL if (import_cols and c <= import_cols) else \
            (HELPER_FILL if import_cols else HDR_FILL)
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        maxlen = max([len(str(ws.cell(row=r, column=c).value or ""))
                      for r in range(1, min(ws.max_row, 200) + 1)] + [10])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 12), 55)
    for r in range(2, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
        if conf_col:
            cell = ws.cell(row=r, column=conf_col)
            try:
                v = float(cell.value)
                cell.fill = HIGH_FILL if v >= 90 else MID_FILL if v >= 75 else LOW_FILL
                cell.alignment = Alignment(horizontal="center")
            except (TypeError, ValueError):
                pass


def _add_instructions(wb):
    ws = wb.create_sheet("تعليمات الاستخدام")
    ws.sheet_view.rightToLeft = True
    lines = [
        ("كيفية استخدام هذا الملف", True),
        ("", False),
        ("الأعمدة الستة الأولى (بالأخضر) في كل ورقة مطابقة تمامًا لأعمدة ملف "
         "الاعتماد price_list — انسخها كما هي والصقها في الملف المعتمد.", False),
        ("الأعمدة الرمادية مساعدة للمراجع فقط (لا تُنسخ): سبب النتيجة، الثقة، "
         "الاقتراح، الطريقة، المصدر.", False),
        ("", False),
        ("ورقة «النتيجة Result»: مطابقات تلقائية فوق حد الثقة — جاهزة للنسخ.", False),
        ("عمود «سبب النتيجة» يخبرك بنظرة لماذا حصلت الخدمة على هذا التصنيف وهل "
         "تحتاج تدقيقًا (مثلًا: «⚠ ثقة حدّية» تعني راجِعها).", False),
        ("", False),
        ("ورقة «تحتاج مراجعة»: لكل صف عمود «اعتماد المراجع» فيه قائمة منسدلة "
         "(نعم/لا/تعديل). صحّح الكود/التصنيف في مكانه إن لزم، ثم اختر «نعم».", False),
        ("• كل خدمة هنا — حتى غير المطابَقة — تحصل على تصنيف رئيسي وفرعي "
         "تقديري بالخبرة الطبية (تحليل/أشعة/أسنان/جراحة/ولادة...)، تجده في "
         "العمودين D و E جاهزًا للتأكيد. يبقى عمود الكود فارغًا ليُكمله المراجع.", False),
        ("ورقة «جاهز للنقل ✅»: تتجمّع فيها تلقائيًا كل الصفوف التي اخترت لها "
         "«نعم» — بأعمدة ملف الاعتماد الستة. انسخها والصقها في الملف المعتمد.", False),
        ("ملاحظة: لو ظهرت #NAME? في ورقة «جاهز للنقل» فإصدار Excel لديك قديم "
         "لا يدعم FILTER — عندها رشّح عمود الاعتماد على «نعم» وانسخ يدويًا.", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(name="Arial", size=13 if bold else 10, bold=bold,
                      color="1F4E78" if bold else "000000")
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 95


def write_output(output_path, matched_df, review_df, summary):
    matched_df.to_excel(output_path, sheet_name="النتيجة Result", index=False)
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as xl:
        review_df.to_excel(xl, sheet_name="تحتاج مراجعة Need Review", index=False)
        summary.to_excel(xl, sheet_name="ملخص Summary", index=False)

    wb = load_workbook(output_path)
    review_sheet = "تحتاج مراجعة Need Review"

    # ورقة النتيجة
    ws1 = wb["النتيجة Result"]
    _style_sheet(ws1, matched_df.shape[1], import_cols=6,
                 conf_col=list(matched_df.columns).index("confidence_score / درجة الثقة") + 1)

    # ورقة المراجعة + قائمة الاعتماد المنسدلة
    ws2 = wb[review_sheet]
    n_rev = review_df.shape[0]
    if n_rev > 0:
        conf_idx = list(review_df.columns).index("confidence_score / درجة الثقة") + 1
        _style_sheet(ws2, review_df.shape[1], import_cols=6, conf_col=conf_idx)
        approve_idx = list(review_df.columns).index(COL_APPROVE) + 1
        approve_letter = get_column_letter(approve_idx)
        dv = DataValidation(type="list", formula1='"نعم,لا,تعديل"', allow_blank=True)
        dv.error = "اختر: نعم أو لا أو تعديل"
        dv.prompt = "اختر «نعم» لنقل الصف إلى ورقة (جاهز للنقل)"
        ws2.add_data_validation(dv)
        dv.add(f"{approve_letter}2:{approve_letter}{n_rev + 1}")
        # فلتر تلقائي: يتيح ترشيح عمود الاعتماد على «نعم» في أي نسخة Excel
        ws2.auto_filter.ref = f"A1:{get_column_letter(review_df.shape[1])}{n_rev + 1}"
    else:
        _style_sheet(ws2, review_df.shape[1], import_cols=6)
        approve_letter, n_rev = "G", 0

    # ورقة «جاهز للنقل» — تتجمّع فيها الصفوف المعتمدة تلقائيًا (FILTER)
    ws3 = wb.create_sheet("جاهز للنقل ✅")
    notes = [
        "ثلاث طرق لنقل الصفوف المعتمدة (اختر الأنسب لنسختك):",
        "(أ) الأضمن في كل النسخ: في ورقة «تحتاج مراجعة» اضغط سهم الفلتر على عمود "
        "«اعتماد المراجع» واختر «نعم»، ثم ظلّل الأعمدة الستة الأولى للصفوف الظاهرة "
        "وانسخها والصقها في ملف الاعتماد.",
        "(ب) عبر أداة التجميع (موصى بها للكميات الكبيرة): بعد اختيار «نعم» احفظ "
        "الملف ثم شغّل: python collect_approved.py \"اسم_الملف.xlsx\" — تُنتج ملفًا "
        "نهائيًا واحدًا بكل الصفوف المعتمدة جاهزًا للّصق.",
        "(ج) تلقائيًا أدناه (يتطلب Excel 365/2021): تظهر الصفوف المعتمدة هنا فور "
        "اختيار «نعم». إن ظهر #NAME? فنسختك أقدم — استخدم الطريقة (أ) أو (ب).",
        "",
    ]
    for i, txt in enumerate(notes, 1):
        c = ws3.cell(row=i, column=1, value=txt)
        c.font = Font(name="Arial", size=10, bold=(i == 1), color="1F4E78")
        c.alignment = Alignment(horizontal="right", wrap_text=True)
    hdr_row = len(notes) + 1
    for j, col in enumerate(STD_COLS, 1):
        ws3.cell(row=hdr_row, column=j, value=col)
    if n_rev > 0:
        last = n_rev + 1
        rng = f"'{review_sheet}'!A2:F{last}"
        cond = f"'{review_sheet}'!{approve_letter}2:{approve_letter}{last}=\"نعم\""
        ws3.cell(row=hdr_row + 1, column=1).value = (
            f'=IFERROR(FILTER({rng},{cond}),"لا توجد صفوف معتمدة بعد — '
            f'أو نسخة Excel لديك لا تدعم FILTER، استخدم الطريقة (أ) أو (ب)")')
    else:
        ws3.cell(row=hdr_row + 1, column=1, value="لا توجد صفوف تحتاج مراجعة.")
    # تنسيق صف العناوين
    for c in range(1, len(STD_COLS) + 1):
        cell = ws3.cell(row=hdr_row, column=c)
        cell.fill = IMPORT_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(c)].width = 24
    ws3.column_dimensions["A"].width = 38
    ws3.sheet_view.rightToLeft = True

    _style_sheet(wb["ملخص Summary"], summary.shape[1])
    _add_instructions(wb)
    # رتّب الأوراق: النتيجة ← المراجعة ← جاهز للنقل ← الملخص ← التعليمات
    order = ["النتيجة Result", review_sheet, "جاهز للنقل ✅",
             "ملخص Summary", "تعليمات الاستخدام"]
    wb._sheets.sort(key=lambda s: order.index(s.title)
                    if s.title in order else 99)
    wb.save(output_path)


# ===========================================================================
# 8) المعالجة الرئيسية  (Main pipeline)
# ===========================================================================
def process(reference, input_file, output, synonyms_path, threshold,
            code_prefix="NEW"):
    print(f"• محرك المطابقة التقريبية: {_FUZZ_ENGINE}")
    syn_map = load_synonyms(synonyms_path)
    print(f"• مرادفات محمّلة: {len(syn_map)} مدخلًا")

    df_ref, exact, exact_syn, choices, choice_list = build_reference(reference, syn_map)
    ref_lookup = choices  # name_norm -> name_norm ; نحتاج خريطة norm->record
    # ابنِ خريطة norm -> record من exact + exact_syn
    ref_record = {}
    ref_record.update(exact)
    for k, v in exact_syn.items():
        ref_record.setdefault(k, v)
    print(f"• خدمات المرجع: {len(df_ref)} صفًا ({len(exact)} اسمًا فريدًا)")
    fidx = FuzzyIndex(choice_list)   # فهرس ضبابي يُبنى مرة واحدة
    catclf = CategoryClassifier(df_ref)   # مُصنِّف طبي يتعلّم من المرجع
    print(f"• المصنّف الطبي: {len(catclf.sub_by_key)} تصنيفًا فرعيًا "
          f"(افتراضي: {catclf.default_key})")

    df_in = ingest.load_records(input_file)
    srcs = df_in["source"].nunique() if "source" in df_in.columns else 1
    print(f"• تمت قراءة الإدخال ({srcs} مصدرًا/ورقة) — خدمات واردة: {len(df_in)}")

    matched_rows, review_rows = [], []
    gen_counter = [0]

    def resolve_code(*candidates):
        """يعيد أول كود غير فارغ، وإلا يولّد كودًا جديدًا في عمود الكود."""
        for c in candidates:
            c = clean_cell(c)
            if c:
                return c
        gen_counter[0] += 1
        return f"{code_prefix}{gen_counter[0]:05d}"

    for _, row in df_in.iterrows():
        name = clean_cell(row["name"])
        if not name:
            continue
        price = "" if pd.isna(row["price"]) else row["price"]
        source = row.get("source", "")
        src_code = clean_cell(row.get("code", ""))   # كود المستشفى الأصلي
        # جرّب الاسم الأساسي والاسم الثانوي (مثل العربي/الإنجليزي) وخذ الأفضل
        res = match_service(name, syn_map, exact, exact_syn, fidx, ref_record)
        name_alt = clean_cell(row.get("name_alt", ""))
        if name_alt and name_alt.lower() != name.lower():
            res_alt = match_service(name_alt, syn_map, exact, exact_syn,
                                    fidx, ref_record)
            if res_alt["score"] > res["score"]:
                res = res_alt
        # اعرض الاسمين معًا في عمود الاسم إن اختلفا (أوضح للمراجع)
        display_name = name if not name_alt or name_alt.lower() == name.lower() \
            else f"{name} | {name_alt}"
        if res["score"] >= threshold:
            # مطابَقة: كود المرجع (الكود المعتمد) ← كود المصدر ← كود مولّد
            matched_rows.append({
                COL_NAME: display_name,
                COL_CODE: resolve_code(res["code"], src_code),
                COL_PRICE: price,
                COL_MAIN: res["main"],
                COL_SUB: res["sub"],
                COL_NOTE: "",
                COL_REASON: build_reason(res, threshold, is_review=False),
                "confidence_score / درجة الثقة": res["score"],
                "match_method / طريقة المطابقة": res["method"],
                "matched_reference / الخدمة المطابقة بالمرجع": res["matched"],
                "source / المصدر": source,
            })
        else:
            # ورقة المراجعة: أول 6 أعمدة بترتيب ملف الاعتماد تمامًا
            # تصنيف طبي تقديري + كود (كود المصدر إن وُجد، وإلا كود مولّد)
            cmain, csub, ckey = catclf.classify(f"{name} {name_alt}")
            review_rows.append({
                COL_NAME: display_name,
                COL_CODE: resolve_code(src_code),
                COL_PRICE: price,
                COL_MAIN: cmain,
                COL_SUB: csub,
                COL_NOTE: "",
                COL_APPROVE: "",
                COL_REASON: build_reason(res, threshold, is_review=True)
                + (f" | تصنيف تقديري: {ckey}" if ckey else ""),
                "confidence_score / درجة الثقة": res["score"],
                "best_suggestion / أقرب اقتراح بالمرجع": res["matched"],
                "match_method / طريقة المطابقة": res["method"],
                "source / المصدر": source,
            })

    matched_df = pd.DataFrame(matched_rows, columns=[
        COL_NAME, COL_CODE, COL_PRICE, COL_MAIN, COL_SUB, COL_NOTE, COL_REASON,
        "confidence_score / درجة الثقة",
        "match_method / طريقة المطابقة",
        "matched_reference / الخدمة المطابقة بالمرجع",
        "source / المصدر"])
    review_df = pd.DataFrame(review_rows, columns=[
        COL_NAME, COL_CODE, COL_PRICE, COL_MAIN, COL_SUB, COL_NOTE,
        COL_APPROVE, COL_REASON, "confidence_score / درجة الثقة",
        "best_suggestion / أقرب اقتراح بالمرجع",
        "match_method / طريقة المطابقة", "source / المصدر"])

    # رتّب النتيجة حسب التصنيف الرئيسي ثم الفرعي ثم الاسم
    if not matched_df.empty:
        matched_df = matched_df.sort_values(
            by=[COL_MAIN, COL_SUB, COL_NAME]).reset_index(drop=True)

    total = len(matched_df) + len(review_df)
    n_match = len(matched_df)
    n_review = len(review_df)
    summary = pd.DataFrame({
        "البند": ["إجمالي الخدمات الواردة", "تمت مطابقتها (Result)",
                  "تحتاج مراجعة (Need Review)", "نسبة المطابقة",
                  "حد القبول (Threshold)", "محرك المطابقة"],
        "القيمة": [total, n_match, n_review,
                   f"{(n_match/total*100 if total else 0):.1f}%",
                   threshold, _FUZZ_ENGINE],
    })

    write_output(output, matched_df, review_df, summary)
    print(f"\n✓ تم الإنشاء: {output}")
    print(f"  - مطابَقة: {n_match}  |  تحتاج مراجعة: {n_review}  |  الإجمالي: {total}")
    return output


def main():
    p = argparse.ArgumentParser(
        description="مُرتِّب ومُطابِق قوائم الأسعار لنظام TPA")
    p.add_argument("--reference", required=True, help="ملف المرجع (الكتالوج الرسمي)")
    p.add_argument("--input", required=True, help="قائمة الأسعار الواردة (فوضوية)")
    p.add_argument("--output", required=True, help="ملف الإخراج المرتّب")
    p.add_argument("--synonyms", default=None, help="ملف قاموس المرادفات JSON (اختياري)")
    p.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD,
                   help=f"حد قبول المطابقة 0-100 (افتراضي {DEFAULT_THRESHOLD})")
    p.add_argument("--code-prefix", default="NEW",
                   help="بادئة الكود المولّد للخدمات بلا كود (افتراضي NEW)")
    args = p.parse_args()
    process(args.reference, args.input, args.output, args.synonyms,
            args.threshold, args.code_prefix)


if __name__ == "__main__":
    main()
