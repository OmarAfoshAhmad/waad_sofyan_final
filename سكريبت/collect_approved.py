#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  collect_approved.py  |  أداة تجميع الصفوف المعتمدة
=============================================================================
تعمل في كل نسخ Excel (لا تعتمد على دالة FILTER).

الفكرة: بعد أن تفتح ملف النتيجة، وتراجع ورقة «تحتاج مراجعة»، وتختار «نعم» في
عمود «اعتماد المراجع» للصفوف التي تريد اعتمادها، **احفظ الملف**، ثم شغّل:

    python collect_approved.py "نتيجة_الخدمات.xlsx"

تُنتج الأداة ملفًا جديدًا واحدًا اسمه «..._للاعتماد_النهائي.xlsx» يحتوي ورقة
واحدة بالأعمدة الستة لملف الاعتماد فقط، تجمع:
   (1) كل صفوف ورقة «النتيجة» (المطابَقة تلقائيًا)
   (2) كل صفوف «تحتاج مراجعة» التي اخترت لها «نعم»
جاهزة لنسخها ولصقها مباشرة في ملف الـ price_list المعتمد.
=============================================================================
"""
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULT_SHEET = "النتيجة Result"
REVIEW_SHEET = "تحتاج مراجعة Need Review"
APPROVE_HINT = "اعتماد"          # نتعرّف على عمود الاعتماد بهذه الكلمة
YES_VALUES = {"نعم", "نعم ", "yes", "y", "true", "1", "✓"}
STD_COLS = [
    "service_name / اسم الخدمة ★",
    "service_code / الكود",
    "contract_price / سعر العقد",
    "main_category / التصنيف الرئيسي",
    "sub_category / البند (التصنيف الفرعي)",
    "notes / ملاحظات",
]


def _first6(df):
    """يأخذ أول 6 أعمدة ويعيد تسميتها بأسماء ملف الاعتماد القياسية."""
    out = df.iloc[:, :6].copy()
    out.columns = STD_COLS
    return out


def collect(path):
    if not os.path.exists(path):
        sys.exit(f"الملف غير موجود: {path}")
    xl = pd.ExcelFile(path)

    # 1) صفوف النتيجة (المطابَقة تلقائيًا)
    result_rows = pd.DataFrame(columns=STD_COLS)
    if RESULT_SHEET in xl.sheet_names:
        result_rows = _first6(pd.read_excel(xl, sheet_name=RESULT_SHEET, dtype=str))

    # 2) صفوف المراجعة المعتمدة (اعتماد = نعم)
    approved_rows = pd.DataFrame(columns=STD_COLS)
    if REVIEW_SHEET in xl.sheet_names:
        rev = pd.read_excel(xl, sheet_name=REVIEW_SHEET, dtype=str)
        appr_col = next((c for c in rev.columns if APPROVE_HINT in str(c)), None)
        if appr_col is None:
            print("تحذير: لم أجد عمود «اعتماد المراجع» في ورقة المراجعة.")
        else:
            mask = rev[appr_col].astype(str).str.strip().str.lower().isin(
                {v.strip().lower() for v in YES_VALUES})
            approved_rows = _first6(rev[mask])
            print(f"• صفوف معتمدة من المراجعة (نعم): {len(approved_rows)}")
    print(f"• صفوف النتيجة التلقائية: {len(result_rows)}")

    final = pd.concat([result_rows, approved_rows], ignore_index=True)
    final = final.fillna("")
    print(f"• الإجمالي النهائي الجاهز للّصق: {len(final)}")

    # 3) اكتب ملفًا نهائيًا بورقة واحدة منسّقة
    base = os.path.splitext(path)[0]
    out_path = f"{base}_للاعتماد_النهائي.xlsx"
    final.to_excel(out_path, sheet_name="النتيجة النهائية", index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    hdr_fill = PatternFill("solid", fgColor="1F6E43")
    for c in range(1, len(STD_COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = get_column_letter(c)
        maxlen = max([len(str(ws.cell(row=r, column=c).value or ""))
                      for r in range(1, min(ws.max_row, 200) + 1)] + [12])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 14), 50)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(STD_COLS) + 1):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
    wb.save(out_path)
    print(f"\n✓ تم الإنشاء: {out_path}")
    print("  افتحه، انسخ كل الجدول، والصقه في ملف الـ price_list المعتمد.")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit('الاستخدام: python collect_approved.py "ملف_النتيجة.xlsx"')
    collect(sys.argv[1])
