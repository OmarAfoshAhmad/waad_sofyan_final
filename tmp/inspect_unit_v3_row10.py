from openpyxl import load_workbook
from pathlib import Path
p=Path(r"D:\tba_waad_system-main_success\tba_waad_system-main\tmp\generated_coverage_imports\استيراد_منافع_مصرف الوحدة_V3.xlsx")
wb=load_workbook(p,data_only=True)
for s in wb.sheetnames:
    ws=wb[s]
    print('\nSHEET',s, ws.max_row, ws.max_column)
    print('headers:', [ws.cell(1,c).value for c in range(1,ws.max_column+1)])
    if ws.max_row >= 10:
        print('row10:', [ws.cell(10,c).value for c in range(1,ws.max_column+1)])
    for r in range(2, ws.max_row+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if 1 in vals and s in ('المنافع','المجموعات'):
            print('row',r,vals)
