from pathlib import Path
import json
import re
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

EXTRACTED = Path("tmp/coverage_extracted")
OUT = Path("tmp/generated_coverage_imports")
OUT.mkdir(parents=True, exist_ok=True)

CATEGORIES = [
    ("CAT-AMBULANCE", "الاسعاف المحلي"),
    ("CAT-ANESTHESIA", "نفقات التخدير"),
    ("CAT-CARDIAC-SURGERY", "عمليات القلب والشرايين"),
    ("CAT-CCU", "عناية القلب"),
    ("CAT-DAY-CARE", "العلاج والرعاية اليومية"),
    ("CAT-DENT-EMERG", "علاج الاسنان بالطوارئ للمريض داخل مستشفى"),
    ("CAT-DIAGNOSTIC", "الكشوف التشخيصية"),
    ("CAT-DRUG", "الدواء"),
    ("CAT-DRUG-CHRONIC", "أدوية الأمراض المزمنة"),
    ("CAT-HOME-NURSING", "التمريض في المنزل أو النقاهة ( بديل الاقامة بعد الخروج )"),
    ("CAT-ICU", "العناية الفائقة"),
    ("CAT-IMG-ADV", "التصوير بالرنين المغناطيسي و المقطعي و الطبقي"),
    ("CAT-IMG-DIAG", "اشعة سينية و اشعة تشخيصية"),
    ("CAT-LAB", "تحاليل و مختبرات"),
    ("CAT-MED-SUP", "المستلزمات الطبية"),
    ("CAT-MAT-COMP", "مضاعفات الحمل والولادة"),
    ("CAT-MAT-CS", "الولادة القيصرية"),
    ("CAT-MAT-NORMAL", "الولادة الطبيعية"),
    ("CAT-DME", "الاجهزه و المعدات الطبية و فق تقرير الطبيب المختص"),
    ("CAT-DENT-ROUTINE", "علاج الاسنان الروتيني ( كشف- خلع- حشو- تنظيف )"),
    ("CAT-DENT-PROSTHO", "علاج الاسنان ( تركيب )"),
    ("CAT-DENT-ORTHO", "علاج الاسنان ( تقويم )"),
    ("CAT-DENT-IMPLANT", "علاج الاسنان ( زراعة )"),
    ("CAT-EYE-EXAM", "كشوف العيون"),
    ("CAT-OPT", "نظارة طبية"),
    ("CAT-ONCOLOGY", "علاج الاورام"),
    ("CAT-PHYSIO", "العلاج الطبيعي"),
    ("CAT-PRACT-FEE", "رسوم الأطباء والجراحين والمستشارين والممارسين"),
    ("CAT-PSYCH-DRUG", "الطب النفسي ( أدوية )"),
    ("CAT-PSYCH-SESS", "الطب النفسي ( جلسات )"),
    ("CAT-ROOM", "الإيواء غرفة خاصة"),
    ("CAT-SURG-MAT", "المعدات والمواد الجراحية"),
    ("CAT-SURGERY", "العمليات الجراحية"),
    ("CAT-TRANSPLANT", "زرع الاعضاء"),
    ("CAT-DIALYSIS", "الغسيل الكلوي"),
]

CATEGORY_CONTEXTS = {
    "CAT-AMBULANCE": ("EMERGENCY",),
    "CAT-HOME-NURSING": ("SPECIAL",),
    "CAT-ROOM": ("INPATIENT",),
    "CAT-ICU": ("INPATIENT",),
    "CAT-CCU": ("INPATIENT",),
    "CAT-DAY-CARE": ("INPATIENT",),
    "CAT-DENT-EMERG": ("INPATIENT",),
    "CAT-DENT-ROUTINE": ("OUTPATIENT",),
    "CAT-DENT-PROSTHO": ("OUTPATIENT",),
    "CAT-DENT-ORTHO": ("OUTPATIENT",),
    "CAT-DENT-IMPLANT": ("OUTPATIENT",),
    "CAT-DIALYSIS": ("INPATIENT", "OUTPATIENT"),
    "CAT-DME": ("OUTPATIENT",),
    "CAT-EYE-EXAM": ("OUTPATIENT",),
    "CAT-MAT-COMP": ("INPATIENT",),
    "CAT-MAT-CS": ("INPATIENT",),
    "CAT-MAT-NORMAL": ("INPATIENT",),
    "CAT-ONCOLOGY": ("INPATIENT", "OUTPATIENT"),
    "CAT-OPT": ("OUTPATIENT",),
    "CAT-PSYCH-DRUG": ("INPATIENT",),
    "CAT-PSYCH-SESS": ("INPATIENT", "OUTPATIENT"),
    "CAT-ANESTHESIA": ("INPATIENT", "OUTPATIENT"),
    "CAT-DIAGNOSTIC": ("INPATIENT", "OUTPATIENT"),
    "CAT-DRUG": ("INPATIENT", "OUTPATIENT"),
    "CAT-IMG-ADV": ("INPATIENT", "OUTPATIENT"),
    "CAT-IMG-DIAG": ("INPATIENT", "OUTPATIENT"),
    "CAT-LAB": ("INPATIENT", "OUTPATIENT"),
    "CAT-MED-SUP": ("INPATIENT", "OUTPATIENT"),
    "CAT-PHYSIO": ("INPATIENT", "OUTPATIENT"),
    "CAT-PRACT-FEE": ("INPATIENT", "OUTPATIENT"),
    "CAT-SURG-MAT": ("INPATIENT", "OUTPATIENT"),
    "CAT-SURGERY": ("INPATIENT", "OUTPATIENT"),
    "CAT-TRANSPLANT": ("INPATIENT",),
}

KEYWORDS = [
    ("CAT-AMBULANCE", ["الاسعاف", "فاعسلاا", "ambulance"]),
    ("CAT-HOME-NURSING", ["التمريض", "ضيرمتلا", "النقاهة", "ةهاقنلا", "nursing"]),
    ("CAT-PHYSIO", ["العلاج الطبيعي", "يعيبطلا جلاعلا", "physio"]),
    ("CAT-IMG-ADV", ["الرنين", "نينرلا", "المقطعي", "يعطقملا", "ct", "mri", "pet"]),
    ("CAT-LAB", ["تحاليل", "ليلحت", "مختبر", "pathology"]),
    ("CAT-IMG-DIAG", ["الأشعة", "ةعش", "radiology", "x-ray"]),
    ("CAT-TRANSPLANT", ["زرع الاعضاء", "ءاضعلاا عرز", "transplant"]),
    ("CAT-DRUG-CHRONIC", ["المزمنة", "ةنمزملا", "chronic"]),
    ("CAT-DRUG", ["الدواء", "الأدوية", "الادوية", "ءاودلا", "drugs"]),
    ("CAT-MED-SUP", ["المستلزمات", "تامزلتسم", "medical requirements"]),
    ("CAT-ICU", ["العناية الفائقة", "ةقئافلا ةيانعلا", "intensive care", "icu"]),
    ("CAT-CCU", ["عناية القلب", "بلقلا ةيانع", "ccu"]),
    ("CAT-PRACT-FEE", ["رسوم الأطباء", "رسوم الاخصائيين", "ءابطلأا موسر", "consultants"]),
    ("CAT-SURG-MAT", ["المواد الجراحية", "ةيحارجلا داوملا", "surgical materials"]),
    ("CAT-ANESTHESIA", ["التخدير", "ريدختلا", "anesthesia"]),
    ("CAT-DIAGNOSTIC", ["الكشوف التشخيصية", "ةيصيخشتلا تافوشكلا", "diagnostic"]),
    ("CAT-DAY-CARE", ["الرعاية اليومية", "ةيمويلا ةياعرلا", "day care"]),
    ("CAT-DENT-EMERG", ["علاج الاسنان بالطوارئ", "داخل المستشفى"]),
    ("CAT-DENT-ROUTINE", ["علاج الاسنان الروتيني", "كشف– خلع", "تنظيف لثة"]),
    ("CAT-DENT-PROSTHO", ["تركيب"]),
    ("CAT-DENT-ORTHO", ["تقويم"]),
    ("CAT-DENT-IMPLANT", ["زراعة"]),
    ("CAT-EYE-EXAM", ["كشوفات العيون"]),
    ("CAT-OPT", ["النظارات الطبية", "نظارة واحدة"]),
    ("CAT-DME", ["الاجهزة والمعدات الطبية", "السماعات الطبية"]),
    ("CAT-MAT-COMP", ["مضاعفات الحمل", "complication of pregnancy"]),
    ("CAT-MAT-NORMAL", ["الولادة الطبيعية", "routine maternity"]),
    ("CAT-MAT-CS", ["القيصرية", "delivery"]),
    ("CAT-ONCOLOGY", ["الاورام", "الأورام"]),
    ("CAT-DIALYSIS", ["غسيل الكلوي", "الغسيل الكلوي"]),
    ("CAT-PSYCH-DRUG", ["الطب النفسي", "أدوية وجلسات"]),
    ("CAT-PSYCH-SESS", ["الطب النفسي", "العلاج النفسي", "أدوية وجلسات"]),
    ("CAT-ROOM", ["غرفة خاصة", "ةصاخ ةفرغ", "accommodation"]),
    ("CAT-SURGERY", ["العمليات الجراحية", "surgery"]),
]

HEADER_BENEFITS = [
    "كود التصنيف", "اسم المنفعة", "السياق", "نسبة التغطية", "نسبة التحمل",
    "موافقة مسبقة", "السقف المالي", "حد المرات", "حد الأيام", "الفترة", "طريقة العد", "نشط"
]
HEADER_GROUPS = [
    "كود المجموعة", "اسم المجموعة", "السياق", "أكواد المنافع مفصولة بفاصلة",
    "السقف المالي", "حد المرات", "حد الأيام", "الفترة", "طريقة العد", "نشط"
]


def norm(s):
    s = (s or "").lower()
    s = re.sub(r"[^\w\u0600-\u06ff%]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def cells_from_doc(data):
    chunks = []
    for t in data.get("tables", []):
        rows = t.get("rows", t) if isinstance(t, dict) else t
        for row in rows:
            text = " ".join(c for c in row if c)
            if text:
                chunks.append(text)
    if not chunks:
        for page in data.get("pages", []):
            chunks.append(page.get("text", ""))
        chunks.extend(data.get("paragraphs", []))
    return chunks


def coverage_from_text(text):
    n = norm(text)
    if "full cover" in n or "fullcover" in n:
        return 100, 0
    m = re.search(r"(\d{1,3})\s*%\s*(?:copay|copayment|كرتشم|مشترك|عفد|دفع)", n)
    if m:
        copay = int(m.group(1))
        return max(0, 100 - copay), copay
    m = re.search(r"(?:copay|copayment|كرتشم|مشترك|عفد|دفع)\s*(\d{1,3})\s*%", n)
    if m:
        copay = int(m.group(1))
        return max(0, 100 - copay), copay
    return None, None


def amount_from_text(text):
    cleaned = text.replace(",", "")
    nums = [int(x) for x in re.findall(r"(?<!\d)(\d{3,7})(?!\d)\s*(?:د\.?ل|د ل|lyd)", cleaned, flags=re.IGNORECASE)]
    if not nums:
        return None
    # When a row contains "100 per night up to 1,000", the enforceable cap is the maximum value.
    return max(nums)


def times_from_text(text):
    n = norm(text)
    m = re.search(r"(\d{1,3})\s*(?:زيارة|زياره|visits?)", n, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def match_categories(text):
    n = norm(text)
    matched = []
    for code, kws in KEYWORDS:
        if any(norm(k) in n for k in kws):
            matched.append(code)
    return matched


def context_for(code, section_context):
    allowed = CATEGORY_CONTEXTS.get(code)
    if not allowed:
        return section_context
    if len(allowed) == 1:
        return allowed[0]
    return section_context if section_context in allowed else allowed[0]


def display_name(code):
    return dict(CATEGORIES).get(code, code)


def generate_workbook(source_name, rows, review):
    wb = Workbook()
    ws = wb.active
    ws.title = "المنافع"
    ws.sheet_view.rightToLeft = True
    ws.append(HEADER_BENEFITS)
    groups = wb.create_sheet("المجموعات")
    groups.sheet_view.rightToLeft = True
    groups.append(HEADER_GROUPS)
    meta = wb.create_sheet("مراجعة")
    meta.sheet_view.rightToLeft = True
    meta.append(["النوع", "الملاحظة", "النص الخام"])

    for row in rows:
        ws.append(row)
    for item in review:
        meta.append(item)

    ref = wb.create_sheet("التصنيفات_المرجعية")
    ref.sheet_view.rightToLeft = True
    ref.append(["كود التصنيف", "اسم التصنيف"])
    for code, name in CATEGORIES:
        ref.append([code, name])

    for sheet in [ws, groups, meta, ref]:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F9E83")
            cell.alignment = Alignment(horizontal="center")
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[chr(64 + col)].width = 24

    dv_context = DataValidation(type="list", formula1='"OUTPATIENT,INPATIENT,ANY"', allow_blank=False)
    dv_yes_no = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=False)
    dv_period = DataValidation(type="list", formula1='"POLICY_PERIOD,ANNUAL,PER_VISIT,PER_SERVICE,LIFETIME"', allow_blank=False)
    dv_count = DataValidation(type="list", formula1='"EACH_UNIT,EACH_LINE,PER_DAY,PER_VISIT"', allow_blank=False)
    ws.add_data_validation(dv_context); ws.add_data_validation(dv_yes_no); ws.add_data_validation(dv_period); ws.add_data_validation(dv_count)
    dv_context.add("C2:C1000"); dv_yes_no.add("F2:F1000"); dv_period.add("J2:J1000"); dv_count.add("K2:K1000")

    ws.freeze_panes = "A2"
    groups.freeze_panes = "A2"
    out = OUT / f"استيراد_منافع_{safe_filename(source_name)}.xlsx"
    wb.save(out)
    return out


def safe_filename(name):
    return re.sub(r'[<>:"/\\\\|?*]+', "_", name).replace(".json", "")


def main():
    outputs = []
    for path in sorted(EXTRACTED.glob("*.json")):
        if path.name.startswith("_"):
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        chunks = cells_from_doc(data)
        rows_by_key = {}
        review = []
        section_context = "INPATIENT"
        if not chunks:
            review.append(["BLOCKER", "لم يتم استخراج نص أو جدول؛ الملف يحتاج OCR/إدخال يدوي", ""])
        for chunk in chunks:
            n = norm(chunk)
            if "العيادات الخارجية" in n or "التردد الخارجي" in n or "outpatient" in n:
                section_context = "OUTPATIENT"
            cov, copay = coverage_from_text(chunk)
            amount = amount_from_text(chunk)
            codes = match_categories(chunk)
            if not codes:
                if cov is not None or amount is not None:
                    review.append(["UNMATCHED", "يوجد سقف/نسبة لكن لم يطابق تصنيفاً معتمداً بثقة", chunk[:500]])
                continue
            for code in codes:
                context = context_for(code, section_context)
                key = (code, context)
                prev = rows_by_key.get(key)
                row_cov, row_copay = cov, copay
                if code in {"CAT-DENT-PROSTHO", "CAT-DENT-ORTHO", "CAT-DENT-IMPLANT"} and "50%" in chunk:
                    row_cov, row_copay = 50, 50
                row_times = times_from_text(chunk)
                row = [
                    code,
                    display_name(code),
                    context,
                    row_cov if row_cov is not None else 100,
                    row_copay if row_copay is not None else 0,
                    "لا",
                    amount,
                    row_times,
                    None,
                    "ANNUAL",
                    "EACH_UNIT",
                    "نعم",
                ]
                if prev is None or (prev[6] is None and amount is not None):
                    rows_by_key[key] = row
        rows = list(rows_by_key.values())
        out = generate_workbook(path.stem, rows, review)
        outputs.append({"source": path.name, "rows": len(rows), "review": len(review), "output": str(out)})
    zip_path = OUT / "قوالب_استيراد_المنافع_المولدة.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for xlsx in OUT.glob("*.xlsx"):
            z.write(xlsx, xlsx.name)
    (OUT / "_generation_summary.json").write_text(json.dumps(outputs, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"outputs": outputs, "zip": str(zip_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
