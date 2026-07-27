from openpyxl import load_workbook
from pathlib import Path
p=Path(r"D:\tba_waad_system-main_success\tba_waad_system-main\tmp\generated_coverage_imports\استيراد_منافع_مصرف الوحدة_V3.xlsx")
wb=load_workbook(p,data_only=True)
errors=[]
for sheet in ["المنافع","المجموعات"]:
    ws=wb[sheet]
    headers=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    print(sheet, headers)
    if "قيمة الفترة" not in headers:
        errors.append(f"{sheet}: missing قيمة الفترة")
    period_col=headers.index("الفترة")+1
    value_col=headers.index("قيمة الفترة")+1
    for r in range(2, ws.max_row+1):
        period=ws.cell(r,period_col).value
        val=ws.cell(r,value_col).value
        if period in {"MULTI_YEAR_POLICY","CUSTOM_DAYS","CUSTOM_WEEKS","CUSTOM_MONTHS","CUSTOM_YEARS"} and (val is None or int(val) <= 1):
            errors.append(f"{sheet} row {r}: {period} value={val}")
        if period in {"CUSTOM_YEARS","MULTI_YEAR_POLICY"}:
            print(sheet, r, ws.cell(r,1).value, ws.cell(r,2).value, period, val)
print("errors", errors)
raise SystemExit(1 if errors else 0)
