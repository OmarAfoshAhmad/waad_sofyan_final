from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import json, zipfile

ROOT = Path(r"D:\tba_waad_system-main_success\tba_waad_system-main")
OUT = ROOT / "tmp" / "benefit_rules_closure_v1"
OUT.mkdir(parents=True, exist_ok=True)

categories = [
("CAT-DIAGNOSTIC","الكشوفات الطبية والتشخيصية","OUTPATIENT,INPATIENT"),
("CAT-DRUG-GENERAL","أدوية الصرف العام","OUTPATIENT"),
("CAT-IMG-DIAG","الأشعة والصور التشخيصية","OUTPATIENT,INPATIENT"),
("CAT-IMG-ADV","التصوير بالرنين المغناطيسي والمقطعي والطبقي","OUTPATIENT,INPATIENT"),
("CAT-ENDOSCOPY","المناظير","OUTPATIENT,INPATIENT"),
("CAT-CARDIO-CHECKUP","فحوصات وتخطيطات القلب غير الجراحية","OUTPATIENT"),
("CAT-LAB","التحاليل الطبية والمختبرات","OUTPATIENT,INPATIENT"),
("CAT-SURGERY","العمليات الجراحية العامة","OUTPATIENT,INPATIENT"),
("CAT-CARDIAC-SURGERY","عمليات القلب والشرايين","INPATIENT"),
("CAT-ROOM","الإيواء في غرفة خاصة أو قسم","INPATIENT"),
("CAT-ICU","العناية الفائقة","INPATIENT"),
("CAT-CCU","عناية القلب","INPATIENT"),
("CAT-PHYSIO","العلاج الطبيعي","OUTPATIENT,INPATIENT"),
("CAT-SPEECH-THERAPY","جلسات علاج النطق للأطفال حتى سن 16 عاماً","OUTPATIENT"),
("CAT-FERTILITY-DRUG","أدوية أمراض الخصوبة والعقم","OUTPATIENT"),
("CAT-THERAPEUTIC-INJ","الحقن العلاجية","OUTPATIENT"),
("CAT-OPT","النظارات الطبية","OUTPATIENT"),
("CAT-DENT-PROSTHO","تركيبات الأسنان","OUTPATIENT"),
("CAT-DME","الأجهزة والمعدات الطبية وفق تقرير الطبيب المختص","OUTPATIENT"),
("CAT-DRUG-CHRONIC","أدوية الأمراض المزمنة","OUTPATIENT"),
("CAT-STIMULANT-DRUG","المنشطات والأدوية المرتبطة بها","OUTPATIENT"),
("CAT-EVAC-COMPANION","تكلفة مرافق واحد للشخص الذي تم إخلاؤه طبياً","INPATIENT"),
("CAT-HEARING-AID","السماعات الطبية","OUTPATIENT"),
("CAT-MED-SUP","المستلزمات الطبية","OUTPATIENT,INPATIENT"),
("CAT-PRACT-FEE","رسوم الأطباء والجراحين والمستشارين والممارسين","OUTPATIENT,INPATIENT"),
("CAT-SURG-MAT","المعدات والمواد الجراحية","OUTPATIENT,INPATIENT"),
("CAT-ANESTHESIA","نفقات التخدير","OUTPATIENT,INPATIENT"),
("CAT-DAY-CARE","العلاج والرعاية اليومية","INPATIENT"),
("CAT-DENT-EMERG","علاج الأسنان الطارئ للمريض داخل المستشفى","INPATIENT"),
("CAT-AMBULANCE","الإسعاف المحلي","EMERGENCY"),
("CAT-HOME-NURSING","التمريض المنزلي أو رعاية النقاهة بعد الخروج","SPECIAL"),
("CAT-PSYCH-DRUG","أدوية الطب النفسي","OUTPATIENT,INPATIENT"),
("CAT-PSYCH-SESS","جلسات الطب النفسي","OUTPATIENT,INPATIENT"),
("CAT-ONCOLOGY","علاج الأورام","OUTPATIENT,INPATIENT"),
("CAT-DIALYSIS","الغسيل الكلوي","OUTPATIENT,INPATIENT"),
("CAT-MAT-NORMAL","الولادة الطبيعية","INPATIENT"),
("CAT-MAT-CS","الولادة القيصرية","INPATIENT"),
("CAT-MAT-COMP","مضاعفات الحمل والولادة","OUTPATIENT,INPATIENT"),
("CAT-DENT-ROUTINE","علاج الأسنان الروتيني","OUTPATIENT"),
("CAT-DENT-ORTHO","تقويم الأسنان","OUTPATIENT"),
("CAT-DENT-IMPLANT","زراعة الأسنان","OUTPATIENT"),
("CAT-EYE-EXAM","كشوفات العيون","OUTPATIENT"),
("CAT-TRANSPLANT","زرع الأعضاء","INPATIENT"),
("CAT-WORK-INJURY","علاج إصابات العمل","INPATIENT"),
("CAT-FAMILY-TRAVEL","سفر أحد أفراد العائلة في حالة الإخلاء الطبي","INPATIENT"),
("CAT-MED-EVAC","الإخلاء الطبي","INPATIENT"),
]

periods = ["PER_SERVICE","PER_VISIT","DAILY","WEEKLY","MONTHLY","QUARTERLY","ANNUAL","CUSTOM_DAYS","CUSTOM_WEEKS","CUSTOM_MONTHS","CUSTOM_YEARS","POLICY_PERIOD","LIFETIME"]

# Representative closure scenarios. These are manual claim-entry samples, not financial advice.
jaliana = [
("JAL-001","عيادات خارجية","CAT-DIAGNOSTIC","كشف طبي/استشارة تشخيصية",1,100,75,"لا يوجد سقف مباشر","POLICY_PERIOD",1,"يتأكد أن المنفعة تظهر مغطاة ولا ينسخ السقف العام في حقل سقف المنفعة"),
("JAL-002","عيادات خارجية","CAT-LAB","CBC وتحاليل مخبرية",1,50,75,"لا يوجد سقف مباشر","POLICY_PERIOD",1,"تحليل خارجي، اختبار تصنيف المختبرات"),
("JAL-003","عيادات خارجية","CAT-IMG-DIAG","أشعة سينية تشخيصية",1,80,75,"لا يوجد سقف مباشر","POLICY_PERIOD",1,"أشعة عادية لا تختلط بالرنين/المقطعي"),
("JAL-004","عيادات خارجية","CAT-IMG-ADV","MRI/CT متقدم",1,900,100,"1500 د.ل", "ANNUAL",1,"سقف مباشر؛ جرب خدمة بسعر 900 ثم كرر حتى يظهر المتبقي"),
("JAL-005","عيادات خارجية","CAT-PHYSIO","جلسة علاج طبيعي",6,100,100,"20 مرة", "ANNUAL",1,"اختبار حد المرات دون سقف مالي مباشر"),
("JAL-006","عيادات خارجية","CAT-SURGERY","عملية صغرى خارجية",1,20000,100,"لا يوجد سقف مباشر", "POLICY_PERIOD",1,"اختبار تأثرها بسقف الوثيقة العام فقط"),
("JAL-007","عيادات خارجية","CAT-DENT-IMPLANT","زراعة سن واحد",1,3000,50,"لا يوجد سقف مباشر", "ANNUAL",1,"اختبار أن زراعة الأسنان مغطاة 50% ولا يظهر سقف عام كمنفعة"),
("JAL-008","إيواء","CAT-ROOM","إقامة غرفة خاصة",3,600,75,"سقف حسب الوثيقة", "ANNUAL",1,"اختبار سياق إيواء فقط"),
("JAL-009","إيواء","CAT-ICU","عناية فائقة يومية",2,1500,75,"سقف/أيام حسب الوثيقة", "ANNUAL",1,"اختبار ICU"),
("JAL-010","إيواء","CAT-CARDIAC-SURGERY","عملية قلب وشرايين",1,25000,75,"سقف حسب الوثيقة", "ANNUAL",1,"تمييز عن الجراحة العامة"),
("JAL-011","إيواء","CAT-MAT-NORMAL","ولادة طبيعية",1,4000,75,"سقف مستقل إن وجد", "ANNUAL",1,"اختبار الولادة الطبيعية"),
("JAL-012","إيواء","CAT-MAT-CS","ولادة قيصرية",1,4000,75,"سقف مستقل إن وجد", "ANNUAL",1,"اختبار الولادة القيصرية"),
("JAL-013","عيادات خارجية","CAT-OPT","نظارة طبية",1,500,75,"500 د.ل", "CUSTOM_YEARS",2,"اختبار كل سنتين"),
("JAL-014","عيادات خارجية","CAT-DENT-PROSTHO","تركيبات أسنان",1,1000,50,"حسب الوثيقة", "ANNUAL",1,"تمييز تركيبات الأسنان عن الزراعة والتقويم"),
("JAL-015","عيادات خارجية","CAT-DME","جهاز/معدة طبية بتقرير",1,1200,75,"حسب الوثيقة", "CUSTOM_YEARS",5,"اختبار كل خمس سنوات"),
("JAL-016","عيادات خارجية","CAT-DRUG-CHRONIC","دواء مرض مزمن",1,300,75,"سقف شهري/سنوي حسب الوثيقة", "MONTHLY",1,"اختبار سقف شهري"),
("JAL-017","عيادات خارجية","CAT-HEARING-AID","سماعة طبية",1,1500,75,"حسب الوثيقة", "CUSTOM_YEARS",5,"اختبار السماعات الطبية"),
("JAL-018","إيواء","CAT-TRANSPLANT","زرع أعضاء",1,25000,75,"حسب الوثيقة", "ANNUAL",1,"اختبار بند إيواء كبير"),
("JAL-019","إيواء","CAT-WORK-INJURY","علاج إصابة عمل",1,25000,75,"حسب الوثيقة", "ANNUAL",1,"اختبار إصابات العمل"),
("JAL-020","دفعة مختلطة","CAT-SURGERY+CAT-IMG-ADV+CAT-PHYSIO","سيناريو تجاوز سقف عام",1,101500,100,"سقف الوثيقة 60000 عند جليانة إن كان مستورداً", "ANNUAL",1,"يدوي: أدخل 3 بنود كما في لقطة الشاشة وتوقع رفض جزئي عند تجاوز السقف العام"),
]

unit = [
("UNIT-001","عيادات خارجية","CAT-DIAGNOSTIC","كشوفات طبية وأدوية الصرف العام",1,600,100,"600 د.ل", "ANNUAL",1,"من وثيقة مصرف الوحدة: كشف/أدوية الصرف العام"),
("UNIT-002","عيادات خارجية","CAT-IMG-DIAG","أشعة وصور ومناظير",1,300,100,"تغطية كاملة/حسب الوثيقة", "ANNUAL",1,"يفضل اختبار IMG-DIAG ثم ENDOSCOPY منفصل"),
("UNIT-003","عيادات خارجية","CAT-ENDOSCOPY","منظار تشخيصي",1,800,100,"حسب الوثيقة", "ANNUAL",1,"تأكد أن المناظير موجودة ولا تسقط إلى أشعة عامة"),
("UNIT-004","عيادات خارجية","CAT-LAB","تحاليل طبية",1,120,100,"تغطية كاملة", "ANNUAL",1,"اختبار المختبرات"),
("UNIT-005","عيادات خارجية","CAT-SURGERY","عملية جراحية صغرى خارجية",1,1000,100,"تغطية كاملة", "ANNUAL",1,"حسب صورة الوحدة: العمليات الجراحية الصغرى خارجية"),
("UNIT-006","إيواء","CAT-SURGERY","عمليات جراحية بأنواعها",1,15000,75,"15000 د.ل", "ANNUAL",1,"اختبار سقف عمليات جراحية في الإيواء"),
("UNIT-007","إيواء","CAT-CARDIAC-SURGERY","عمليات القلب والشرايين",1,20000,75,"20000 د.ل", "ANNUAL",1,"اختبار سقف عمليات القلب"),
("UNIT-008","إيواء","CAT-ROOM","الإيواء أو القسم",1,15000,75,"15000 د.ل", "ANNUAL",1,"اختبار الإيواء"),
("UNIT-009","إيواء","CAT-ICU","الإيواء في العناية",1,30000,75,"30000 د.ل", "ANNUAL",1,"اختبار العناية"),
("UNIT-010","عيادات خارجية","CAT-PHYSIO","جلسة علاج طبيعي",30,50,100,"30 جلسة خلال السنة", "ANNUAL",1,"اختبار حد 30 جلسة"),
("UNIT-011","عيادات خارجية","CAT-SPEECH-THERAPY","جلسة علاج نطق لطفل",70,40,100,"70 جلسة خلال السنة", "ANNUAL",1,"اختبار حد الجلسات للأطفال"),
("UNIT-012","عيادات خارجية","CAT-FERTILITY-DRUG","أدوية الخصوبة والعقم",1,1000,100,"5000 كل 3 سنوات", "CUSTOM_YEARS",3,"اختبار فترة مخصصة 3 سنوات"),
("UNIT-013","عيادات خارجية","CAT-THERAPEUTIC-INJ","حقنة علاجية",4,750,100,"3000 للأسرة الواحدة 4 مرات", "ANNUAL",1,"اختبار حد عدد مرات + سقف مالي"),
("UNIT-014","عيادات خارجية","CAT-OPT","نظارة طبية",1,500,100,"500 د.ل كل سنتين", "CUSTOM_YEARS",2,"اختبار كل سنتين"),
("UNIT-015","عيادات خارجية","CAT-DENT-PROSTHO","تركيبات أسنان",1,2000,80,"80% من قيمة الفاتورة", "ANNUAL",1,"اختبار تحمل/تغطية مختلفة"),
("UNIT-016","عيادات خارجية","CAT-DME","سماعات/معدات طبية",1,5000,100,"5000 كل 5 سنوات", "CUSTOM_YEARS",5,"اختبار DME، والسماعات يفضل CAT-HEARING-AID إذا منفصلة"),
("UNIT-017","عيادات خارجية","CAT-DRUG-CHRONIC","أدوية أمراض مزمنة شهرية",1,300,100,"300 د.ل شهرياً", "MONTHLY",1,"اختبار سقف شهري"),
("UNIT-018","عيادات خارجية","CAT-PSYCH-DRUG","أدوية طب نفسي",1,3000,75,"3000 د.ل", "ANNUAL",1,"اختبار أدوية الطب النفسي"),
("UNIT-019","عيادات خارجية","CAT-PSYCH-SESS","جلسات طب نفسي",25,120,75,"3000 د.ل / 25 جلسة", "ANNUAL",1,"اختبار سقف مشترك محتمل مع psych drug حسب إعدادك"),
("UNIT-020","إيواء","CAT-AMBULANCE","إسعاف محلي",1,600,100,"حسب الوثيقة", "PER_SERVICE",1,"اختبار إسعاف/طوارئ"),
("UNIT-021","إيواء","CAT-MED-EVAC","إخلاء طبي",1,10000,75,"حسب الوثيقة", "ANNUAL",1,"اختبار الإخلاء الطبي"),
("UNIT-022","إيواء","CAT-EVAC-COMPANION","تكلفة مرافق واحد",1,2000,75,"حسب الوثيقة", "ANNUAL",1,"اختبار مرافق الإخلاء"),
("UNIT-023","إيواء","CAT-FAMILY-TRAVEL","سفر أحد أفراد العائلة",1,1500,75,"حسب الوثيقة", "ANNUAL",1,"اختبار سفر العائلة"),
("UNIT-024","عيادات خارجية","CAT-HEARING-AID","سماعة طبية",1,5000,100,"5000 كل 5 سنوات إن فصلتها عن DME", "CUSTOM_YEARS",5,"اختبار التصنيف المنفصل للسماعات"),
]

edge = [
("EDGE-001","أي وثيقة","CAT-IMG-ADV","سقف مالي مباشر",2,900,"أول بند يستهلك، الثاني يرفض جزئياً إذا تجاوز 1500"),
("EDGE-002","أي وثيقة","CAT-PHYSIO","حد مرات",21,100,"المرات بعد الحد ترفض/تظهر متبقي 0"),
("EDGE-003","أي وثيقة","CAT-OPT","كل سنتين",1,500,"مطالبتان في نفس فترة السنتين: الثانية تستهلك من نفس الفترة"),
("EDGE-004","أي وثيقة","CAT-DME","كل خمس سنوات",1,5000,"مطالبة بعد 5 سنوات تبدأ فترة جديدة"),
("EDGE-005","أي وثيقة","CAT-DRUG-CHRONIC","شهري",1,300,"مطالبة في شهر جديد يجب أن يكون لها فترة جديدة"),
("EDGE-006","أي وثيقة","CAT-SURGERY+CAT-ANESTHESIA+CAT-SURG-MAT","مجموعة مشتركة",3,0,"اختبار سقف مشترك للعملية إن أعددته كمجموعة"),
("EDGE-007","أي وثيقة","CAT-DIAGNOSTIC","بلا سقف مباشر",1,100,"لا يظهر السقف العام في حقل سقف المنفعة"),
("EDGE-008","أي وثيقة","CAT-DENT-IMPLANT","غير مغطى في سياق إيواء إذا التصنيف خارجي",1,3000,"اختبر خطأ السياق: INPATIENT يجب لا يغطي إذا غير مسموح"),
]

services = []
for prefix, rows in [("JAL", jaliana), ("UNIT", unit)]:
    for idx, r in enumerate(rows, 1):
        if "+" in r[2]:
            continue
        services.append((f"{prefix}-CLS-{idx:03d}", r[3], r[2], r[1], r[5], r[0]))


def style(ws):
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill("solid", fgColor="2F9E83")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = 24

wb = Workbook()
ws = wb.active
ws.title = "تعليمات"
ws.append(["البند", "التوضيح"])
for row in [
    ("الغرض", "عينات إدخال يدوي لإغلاق قواعد التغطية لجليانة ومصرف الوحدة."),
    ("مهم", "هذه ليست ملف استيراد مطالبات آلي؛ هي سيناريوهات تدخلها يدوياً في شاشة المطالبات بعد استيراد الوثيقة وقائمة الأسعار."),
    ("التحقق", "قارن حصة الشركة/المشترك، حقل سقف المنفعة، المتبقي من السقف، والرفض الجزئي."),
    ("قاعدة ذهبية", "الخدمات بلا سقف مباشر يجب ألا تعرض السقف العام في حقل سقف المنفعة."),
]: ws.append(row)

cat_ws = wb.create_sheet("التصنيفات_الموحدة_V1")
cat_ws.append(["كود التصنيف", "اسم العرض المعتمد", "السياقات المسموحة"])
for r in categories: cat_ws.append(r)

svc_ws = wb.create_sheet("قائمة_خدمات_اختبار")
svc_ws.append(["كود خدمة مقترح", "اسم الخدمة", "كود التصنيف", "السياق", "سعر افتراضي", "سيناريو"])
for r in services: svc_ws.append(r)

for title, rows in [("عينات_جليانة", jaliana), ("عينات_مصرف_الوحدة", unit)]:
    sh = wb.create_sheet(title)
    sh.append(["رقم السيناريو", "السياق", "كود التصنيف", "اسم الخدمة للاختبار", "الكمية", "سعر الوحدة", "نسبة التغطية المتوقعة %", "السقف/القيد المباشر المتوقع", "نوع مدة السقف", "قيمة مدة السقف", "ماذا أراقب؟"])
    for r in rows: sh.append(r)

edge_ws = wb.create_sheet("سيناريوهات_الحواف")
edge_ws.append(["رقم", "الوثيقة", "التصنيف/المجموعة", "نوع الحالة", "الكمية", "سعر الوحدة", "المتوقع"])
for r in edge: edge_ws.append(r)

valid_ws = wb.create_sheet("قيم_مسموحة")
valid_ws.append(["period_type", "متى يستخدم"])
for p in periods:
    valid_ws.append([p, {
        "PER_SERVICE":"كل خدمة مستقلة", "PER_VISIT":"كل زيارة", "DAILY":"يومي", "WEEKLY":"أسبوعي", "MONTHLY":"شهري", "QUARTERLY":"ربع سنوي", "ANNUAL":"سنوي", "CUSTOM_DAYS":"كل عدد أيام", "CUSTOM_WEEKS":"كل عدد أسابيع", "CUSTOM_MONTHS":"كل عدد أشهر", "CUSTOM_YEARS":"كل عدد سنوات", "POLICY_PERIOD":"مدة الوثيقة", "LIFETIME":"مدى الحياة"
    }.get(p, "")])

for sh in wb.worksheets: style(sh)

# validations on sample sheets
for sh in [wb["عينات_جليانة"], wb["عينات_مصرف_الوحدة"]]:
    dv_context = DataValidation(type="list", formula1='"عيادات خارجية,إيواء,طوارئ,خاص,دفعة مختلطة"')
    dv_period = DataValidation(type="list", formula1='"' + ','.join(periods) + '"')
    sh.add_data_validation(dv_context); sh.add_data_validation(dv_period)
    dv_context.add(f"B2:B{sh.max_row}")
    dv_period.add(f"I2:I{sh.max_row}")

out_xlsx = OUT / "عينات_اختبار_إغلاق_قواعد_التغطية_جليانة_ومصرف_الوحدة_V1.xlsx"
wb.save(out_xlsx)

# closure doc
closure = ROOT / "docs" / "BENEFIT_COVERAGE_RULES_CLOSURE_V1.md"
closure.write_text("""# إغلاق قواعد التغطية V1

تاريخ الإغلاق: 2026-07-27

## القرار

اعتماد قائمة تصنيفات موحدة من 46 تصنيفاً لتغطية وثيقتي جليانة ومصرف الوحدة، مع اعتبار السياق وقيمة مدة السقف جزءاً حاكماً من القاعدة وليس وصفاً بصرياً.

## القواعد الحاكمة

1. لا يتم إنشاء تصنيف جديد إلا إذا كان له سقف أو نسبة أو شرط مختلف فعلاً.
2. الخدمات بلا سقف مباشر لا تعرض السقف العام داخل حقل سقف المنفعة.
3. السقف العام يطبق كحارس نهائي على إجمالي المطالبة/الوثيقة، وليس كأنه سقف منفعة.
4. المجموعات تستخدم عندما تشترك عدة منافع في سقف واحد.
5. مدة السقف تتكون من `periodType` و `periodValue`، وتدخل في محرك التغطية ودفتر الاستهلاك.

## أنواع المدة المعتمدة

- PER_SERVICE
- PER_VISIT
- DAILY
- WEEKLY
- MONTHLY
- QUARTERLY
- ANNUAL
- CUSTOM_DAYS
- CUSTOM_WEEKS
- CUSTOM_MONTHS
- CUSTOM_YEARS
- POLICY_PERIOD
- LIFETIME

## التحقق الآلي المنفذ

- Backend compile
- Frontend build
- BenefitBucketLedgerServiceTest
- BenefitStructureServiceTest

## ملف العينات اليدوية

`tmp/benefit_rules_closure_v1/عينات_اختبار_إغلاق_قواعد_التغطية_جليانة_ومصرف_الوحدة_V1.xlsx`

استخدمه بعد استيراد الوثائق وقوائم الأسعار لاختبار الحالات اليدوية.
""", encoding="utf-8")

summary = {
    "categories": len(categories),
    "jaliana_scenarios": len(jaliana),
    "unit_bank_scenarios": len(unit),
    "edge_scenarios": len(edge),
    "files": [str(out_xlsx), str(closure)]
}
(OUT / "closure_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
zip_path = OUT / "حزمة_إغلاق_قواعد_التغطية_V1.zip"
with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
    z.write(out_xlsx, out_xlsx.name)
    z.write(closure, closure.name)
    z.write(OUT / "closure_summary.json", "closure_summary.json")
print(json.dumps({**summary, "zip": str(zip_path)}, ensure_ascii=False, indent=2))
