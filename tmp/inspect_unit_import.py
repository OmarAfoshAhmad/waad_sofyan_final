from openpyxl import load_workbook
from pathlib import Path
p=Path(r"D:\tba_waad_system-main_success\tba_waad_system-main\tmp\generated_coverage_imports\استيراد_منافع_مصرف الوحدة_V2.xlsx")
wb=load_workbook(p,data_only=True)
for wsname in wb.sheetnames:
    print(wsname, wb[wsname].max_row, wb[wsname].max_column)
for sheet in ['Buckets','المنافع','المجموعات']:
    if sheet in wb.sheetnames:
        ws=wb[sheet]
        print('\nSHEET',sheet)
        print([ws.cell(1,c).value for c in range(1, ws.max_column+1)])
        for r in range(1, min(ws.max_row,25)+1):
            vals=[ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            if r in [1,17,19,21] or any(v in ('MULTI_YEAR_POLICY','CUSTOM_YEARS') for v in vals):
                print(r, vals)
