from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

src=Path(r"D:\tba_waad_system-main_success\tba_waad_system-main\tmp\generated_coverage_imports\استيراد_منافع_مصرف الوحدة_V2.xlsx")
out=src.with_name("استيراد_منافع_مصرف الوحدة_V3.xlsx")
wb=load_workbook(src)

# Benefits simplified sheet: insert period value after period column (J -> K)
ws=wb["المنافع"]
headers=[ws.cell(1,c).value for c in range(1, ws.max_column+1)]
if "قيمة الفترة" not in headers:
    ws.insert_cols(11)
    ws.cell(1,11).value="قيمة الفترة"
    for r in range(2, ws.max_row+1):
        ws.cell(r,11).value=1
# Set exact multi-year values and prefer new CUSTOM_YEARS enum
for r in range(2, ws.max_row+1):
    code=ws.cell(r,1).value
    period=ws.cell(r,10).value
    if code == "CAT-FERTILITY-DRUG":
        ws.cell(r,10).value="CUSTOM_YEARS"
        ws.cell(r,11).value=3
    elif code == "CAT-OPT":
        ws.cell(r,10).value="CUSTOM_YEARS"
        ws.cell(r,11).value=2
    elif code in {"CAT-HEARING-AID", "CAT-DME"} and period in {"MULTI_YEAR_POLICY", "CUSTOM_YEARS"}:
        ws.cell(r,10).value="CUSTOM_YEARS"
        ws.cell(r,11).value=5
    elif period in {"MULTI_YEAR_POLICY", "CUSTOM_YEARS"} and (ws.cell(r,11).value in (None, "", 1)):
        raise SystemExit(f"Row {r} has multi-year period without explicit value: {code}")

# Groups sheet: keep compatible with new template too
if "المجموعات" in wb.sheetnames:
    gs=wb["المجموعات"]
    gheaders=[gs.cell(1,c).value for c in range(1, gs.max_column+1)]
    if "قيمة الفترة" not in gheaders:
        gs.insert_cols(9)
        gs.cell(1,9).value="قيمة الفترة"
        for r in range(2, gs.max_row+1):
            gs.cell(r,9).value=1

# Add/refresh validations on benefits
for dv in list(ws.data_validations.dataValidation):
    pass
# openpyxl keeps old validations; add new ones to correct columns.
dv_period=DataValidation(type="list", formula1='"PER_SERVICE,PER_VISIT,DAILY,WEEKLY,MONTHLY,QUARTERLY,ANNUAL,CUSTOM_DAYS,CUSTOM_WEEKS,CUSTOM_MONTHS,CUSTOM_YEARS,POLICY_PERIOD,LIFETIME"')
dv_count=DataValidation(type="list", formula1='"EACH_UNIT,EACH_LINE,PER_DAY,PER_VISIT"')
dv_active=DataValidation(type="list", formula1='"نعم,لا"')
ws.add_data_validation(dv_period); dv_period.add(f"J2:J{max(ws.max_row,1000)}")
ws.add_data_validation(dv_count); dv_count.add(f"L2:L{max(ws.max_row,1000)}")
ws.add_data_validation(dv_active); dv_active.add(f"M2:M{max(ws.max_row,1000)}")

# Style headers and sizes
for sheet in wb.worksheets:
    sheet.sheet_view.rightToLeft=True
    for cell in sheet[1]:
        cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="2F9E83")
        cell.alignment=Alignment(horizontal="center")
    for col in range(1, sheet.max_column+1):
        sheet.column_dimensions[__import__('openpyxl').utils.get_column_letter(col)].width=24

wb.save(out)
print(out)
