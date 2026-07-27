from pathlib import Path
from openpyxl import load_workbook
import json, sys
ROOT = Path(r"D:\tba_waad_system-main_success\tba_waad_system-main")
file = ROOT / "tmp" / "benefit_rules_closure_v1" / "عينات_اختبار_إغلاق_قواعد_التغطية_جليانة_ومصرف_الوحدة_V1.xlsx"
allowed_periods = {"PER_SERVICE","PER_VISIT","DAILY","WEEKLY","MONTHLY","QUARTERLY","ANNUAL","CUSTOM_DAYS","CUSTOM_WEEKS","CUSTOM_MONTHS","CUSTOM_YEARS","POLICY_PERIOD","LIFETIME"}
wb = load_workbook(file, data_only=True)
cat_codes = {row[0] for row in wb["التصنيفات_الموحدة_V1"].iter_rows(min_row=2, values_only=True) if row[0]}
errors=[]
for sheet in ["عينات_جليانة", "عينات_مصرف_الوحدة"]:
    ws=wb[sheet]
    for n,row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        scenario, context, code, name, qty, price, coverage, limit, period, period_value, watch = row
        if not scenario: continue
        if code and "+" not in code and code not in cat_codes:
            errors.append(f"{sheet} row {n}: unknown category {code}")
        if period not in allowed_periods:
            errors.append(f"{sheet} row {n}: invalid period {period}")
        if period in {"CUSTOM_DAYS","CUSTOM_WEEKS","CUSTOM_MONTHS","CUSTOM_YEARS"} and (period_value is None or int(period_value) < 2):
            errors.append(f"{sheet} row {n}: custom period needs value >=2")
edge = wb["سيناريوهات_الحواف"].max_row - 1
summary={"categories":len(cat_codes),"jaliana_rows":wb["عينات_جليانة"].max_row-1,"unit_rows":wb["عينات_مصرف_الوحدة"].max_row-1,"edge_rows":edge,"errors":errors}
print(json.dumps(summary, ensure_ascii=False, indent=2))
if errors: sys.exit(1)
