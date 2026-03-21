import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC_PATH = "draft/قائمة اسعار خدمات دار الشفاء مصنفة.xlsx"
OUT_PATH = "draft/دار_الشفاء_جاهز_للاستيراد.xlsx"

# Mapping file main category -> system main category
MAIN_CAT_MAP = {
    "إيواء": "إيواء",
    "عيادات خارجية": "عيادات خارجية",
    "اشعة": "عيادات خارجية",
    "اشعة ": "عيادات خارجية",
    "تحاليل طبية": "عيادات خارجية",
    "علاج طبيعي": "عيادات خارجية",
    "علاج طبيعي ": "عيادات خارجية",
    "عمليات": "إيواء",
    "عمليات ": "إيواء",
    "اسنان تجميلي": "عيادات خارجية",
    "اسنان وقائي": "عيادات خارجية",
    "اسنان وقائي ": "عيادات خارجية",
}

# Mapping file sub category -> system sub category
SUB_CAT_MAP = {
    "اشعة": "أشعة تحاليل رسوم أطباء",
    "اشعة ": "أشعة تحاليل رسوم أطباء",
    "خدمات الأسنان": "أسنان روتيني",
    "خدمات العلاج الطبيعي": "علاج طبيعي",
    "خدمات الرعاية بالعناية المركزه": "عام",
    "خدمات الرعايه الطبيه": "عام",
    "خدمات التخذير": "عام",
    "خدمات الجراحة": "عام",
    "خدمات الاذن والانف والحنجرة": "عام",
    "خدمات الصور التشخيصية": "أشعة تحاليل رسوم أطباء",
    "خدمات الطوارئ": "عام",
    "خدمات العظام": "عام",
    "خدمات العلاج الكيماوي": "عام",
    "خدمات العيادات الخارجية": "عام",
    "خدمات العيون": "عام",
    "خدمات المناظير": "عام",
    "خدمات تخطيط العصب": "عام",
    "خدمات جراحة التجميل": "عام",
    "خدمات جراحة الصدر": "عام",
    "خدمات جلسات الغسيل": "عام",
    "كشف": "عام",
    "مراجعة": "عام",
    "معامل": "أشعة تحاليل رسوم أطباء",
    "التخصص": "",
}

DENTAL_COSMETIC_MAIN = {"اسنان تجميلي"}
DENTAL_ROUTINE_MAIN = {"اسنان وقائي", "اسنان وقائي "}
RADIOLOGY_MAIN = {"اشعة", "اشعة "}
LABS_MAIN = {"تحاليل طبية"}
PHYSIO_MAIN = {"علاج طبيعي", "علاج طبيعي "}

CODE_PATTERN = re.compile(r"[A-Z]{2,4}-[A-Z0-9-]+")


def extract_code(service_name: str) -> str:
    match = CODE_PATTERN.search(service_name or "")
    return match.group() if match else ""


def classify(main_cat: str, sub_cat: str):
    main = (main_cat or "").strip()
    sub = (sub_cat or "").strip()

    system_main = MAIN_CAT_MAP.get(main, main if main else "عيادات خارجية")

    if main in DENTAL_COSMETIC_MAIN:
        system_sub = "أسنان تجميلي"
    elif main in DENTAL_ROUTINE_MAIN:
        system_sub = "أسنان روتيني"
    elif main in RADIOLOGY_MAIN or main in LABS_MAIN:
        system_sub = "أشعة تحاليل رسوم أطباء"
    elif main in PHYSIO_MAIN:
        system_sub = "علاج طبيعي"
    else:
        system_sub = SUB_CAT_MAP.get(sub, sub)

    if not system_sub:
        system_sub = "عام"

    return system_main, system_sub


def read_services(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet"]

    services = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        price = row[2]
        service = row[3]
        sub_cat = row[4]
        main_cat = row[5]

        if service and price and str(price) != "السعر" and str(service) != "الخدمه":
            services.append(
                {
                    "service": str(service).strip(),
                    "price": float(price),
                    "sub_cat": str(sub_cat).strip() if sub_cat else "",
                    "main_cat": str(main_cat).strip() if main_cat else "",
                }
            )

    return services


def build_workbook(services, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing_Template"
    ws.sheet_view.rightToLeft = True

    required_header_fill = PatternFill("solid", fgColor="1F4E79")
    optional_header_fill = PatternFill("solid", fgColor="D9D9D9")
    example_fill = PatternFill("solid", fgColor="FFF2CC")

    required_font = Font(bold=True, color="FFFFFF")
    optional_font = Font(bold=True)
    example_font = Font(italic=True, color="666666")

    header_border = Border(bottom=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    headers = [
        ("service_name / اسم الخدمة ★", True),
        ("service_code / الكود", False),
        ("standard_price / السعر الأساسي", False),
        ("contract_price / سعر العقد", False),
        ("main_category / التصنيف الرئيسي", False),
        ("sub_category / البند (التصنيف الفرعي)", False),
        ("specialty / التخصص", False),
        ("notes / ملاحظات", False),
    ]

    for col_idx, (text, required) in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, text)
        cell.font = required_font if required else optional_font
        cell.fill = required_header_fill if required else optional_header_fill
        cell.border = header_border
        cell.alignment = header_alignment

    example = ["فحص شامل", "MC-001", 120.0, 100.0, "عيادات خارجية", "عام", "باطنة", "مثال - احذف هذا الصف"]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(2, col_idx, value)
        cell.fill = example_fill
        cell.font = example_font

    row_num = 3
    for item in services:
        service_name = item["service"]
        code = extract_code(service_name)
        price = item["price"]
        specialty = item["sub_cat"]
        system_main, system_sub = classify(item["main_cat"], item["sub_cat"])

        row = [service_name, code, price, price, system_main, system_sub, specialty, ""]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_num, col_idx, value)
        row_num += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 40

    inst = wb.create_sheet("التعليمات")
    inst.sheet_view.rightToLeft = True
    title_font = Font(bold=True, size=14)

    r = 1
    inst.cell(r, 1, "معلومات الملف:").font = title_font
    r += 1
    inst.cell(r, 1, "المنشأة: دار الشفاء")
    r += 1
    inst.cell(r, 1, f"عدد الخدمات: {len(services)}")
    r += 2

    inst.cell(r, 1, "تعليمات الاستخدام:").font = title_font
    r += 1
    inst.cell(r, 1, "1. هذا القالب مطابق لقالب استيراد عقود مقدمي الخدمة في النظام")
    r += 1
    inst.cell(r, 1, "2. service_name هو العمود الإلزامي")
    r += 1
    inst.cell(r, 1, "3. تم ملء main_category و sub_category بناء على تصنيفات إدارة التصنيفات")
    r += 1
    inst.cell(r, 1, "4. specialty يحتوي تخصص مقدم الخدمة الأصلي")
    r += 1

    inst.column_dimensions["A"].width = 85

    wb.save(out_path)


def main():
    services = read_services(SRC_PATH)
    build_workbook(services, OUT_PATH)
    print(f"Done. Services: {len(services)}")
    print(f"Saved file: {OUT_PATH}")


if __name__ == "__main__":
    main()
