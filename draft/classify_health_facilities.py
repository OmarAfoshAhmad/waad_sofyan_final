import argparse
import json
import os
import re
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

# System categories (from medical_categories seed)
SYSTEM_MAIN_CATEGORIES = {"إيواء", "عيادات خارجية"}
SYSTEM_SUB_CATEGORIES = {
    "عام",
    "علاج طبيعي",
    "أسنان روتيني",
    "أسنان تجميلي",
    "أشعة تحاليل رسوم أطباء",
    "رنين مغناطيسي",
    "علاجات وأدوية روتينية",
    "أجهزة ومعدات",
    "النظارة الطبية"
}

CODE_PATTERN = re.compile(r"[A-Z]{2,4}-[A-Z0-9-]+")


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def extract_code(service_name):
    match = CODE_PATTERN.search(service_name or "")
    return match.group() if match else ""


def normalize_for_key(value):
    text = normalize(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


class AIClassifier:
    def __init__(self, model, endpoint, api_key, temperature=0.0, timeout=45):
        self.model = model
        self.endpoint = endpoint
        self.api_key = api_key
        self.temperature = temperature
        self.timeout = timeout
        self.enabled = bool(self.api_key and self.model and self.endpoint)
        self.cache = {}

    def classify(self, service_name, main_hint, sub_hint):
        cache_key = (
            normalize_for_key(service_name),
            normalize_for_key(main_hint),
            normalize_for_key(sub_hint)
        )
        if cache_key in self.cache:
            return self.cache[cache_key]

        if not self.enabled:
            result = self._fallback_classification(main_hint, sub_hint)
            self.cache[cache_key] = result
            return result

        try:
            result = self._call_llm(service_name, main_hint, sub_hint)
        except Exception:
            result = self._fallback_classification(main_hint, sub_hint)

        self.cache[cache_key] = result
        return result

    def _fallback_classification(self, main_hint, sub_hint):
        main = normalize(main_hint)
        sub = normalize(sub_hint)

        if main not in SYSTEM_MAIN_CATEGORIES:
            main = "عيادات خارجية"

        if sub not in SYSTEM_SUB_CATEGORIES:
            sub = "عام"

        return {
            "main": main,
            "sub": sub,
            "confidence": 0.2,
            "method": "fallback"
        }

    def _call_llm(self, service_name, main_hint, sub_hint):
        system_prompt = (
            "You are a medical taxonomy classifier for provider contract price lists. "
            "Return only valid JSON with no markdown. "
            "Choose exactly one main_category and one sub_category from the allowed lists."
        )

        user_payload = {
            "task": "classify_medical_service",
            "allowed_main_categories": sorted(SYSTEM_MAIN_CATEGORIES),
            "allowed_sub_categories": sorted(SYSTEM_SUB_CATEGORIES),
            "service_name": normalize(service_name),
            "source_main_hint": normalize(main_hint),
            "source_sub_hint": normalize(sub_hint),
            "output_schema": {
                "main_category": "one of allowed_main_categories",
                "sub_category": "one of allowed_sub_categories",
                "confidence": "number between 0 and 1",
                "reason": "short text"
            }
        }

        body = {
            "model": self.model,
            "temperature": self.temperature,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)}
            ]
        }

        req = urllib.request.Request(
            self.endpoint,
            data=json.dumps(body).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}"
            },
            method="POST"
        )

        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                raw = resp.read().decode("utf-8")
        except urllib.error.HTTPError as http_err:
            details = http_err.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"LLM HTTP {http_err.code}: {details}") from http_err

        parsed = json.loads(raw)
        content = parsed["choices"][0]["message"]["content"]
        model_result = json.loads(content)

        main = normalize(model_result.get("main_category"))
        sub = normalize(model_result.get("sub_category"))
        confidence = model_result.get("confidence", 0.5)

        if main not in SYSTEM_MAIN_CATEGORIES:
            main = "عيادات خارجية"
        if sub not in SYSTEM_SUB_CATEGORIES:
            sub = "عام"

        try:
            confidence_value = float(confidence)
        except Exception:
            confidence_value = 0.5
        confidence_value = max(0.0, min(1.0, confidence_value))

        return {
            "main": main,
            "sub": sub,
            "confidence": confidence_value,
            "method": "ai"
        }


def detect_columns(ws):
    # Try header-based detection first
    for r in range(1, min(ws.max_row, 60) + 1):
        values = [normalize(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if not any(values):
            continue

        service_col = None
        price_col = None
        sub_col = None
        main_col = None

        for idx, val in enumerate(values, start=1):
            if not val:
                continue
            if service_col is None and ("service_name" in val or "اسم الخدمة" in val or "الخدمه" in val or val == "الخدمة"):
                service_col = idx
            if price_col is None and ("contract_price" in val or "unit_price" in val or "price" in val or "السعر" in val):
                price_col = idx
            if sub_col is None and ("sub_category" in val or "التصنيف الفرعي" in val or "التخصص" in val):
                sub_col = idx
            if main_col is None and ("main_category" in val or "التصنيف الرئيسي" in val or "category" == val):
                main_col = idx

        if service_col and price_col:
            if main_col is None:
                # Infer main category column by sampling next rows for known values.
                known_main_values = {
                    "إيواء", "عيادات خارجية", "عمليات", "عمليات ", "اشعة", "اشعة ",
                    "تحاليل طبية", "علاج طبيعي", "علاج طبيعي ", "اسنان تجميلي", "اسنان وقائي", "اسنان وقائي "
                }
                best_col = None
                best_hits = 0
                for c in range(1, ws.max_column + 1):
                    if c in {service_col, price_col, sub_col}:
                        continue
                    hits = 0
                    for rr in range(r + 1, min(ws.max_row, r + 120) + 1):
                        v = normalize(ws.cell(rr, c).value)
                        if v in known_main_values:
                            hits += 1
                    if hits > best_hits:
                        best_hits = hits
                        best_col = c
                if best_col is not None and best_hits >= 3:
                    main_col = best_col

            return {
                "header_row": r,
                "service_col": service_col,
                "price_col": price_col,
                "sub_col": sub_col,
                "main_col": main_col
            }

    # Fallback for known hospital format: C price, D service, E specialty, F main
    return {
        "header_row": 10,
        "service_col": 4,
        "price_col": 3,
        "sub_col": 5,
        "main_col": 6
    }


def extract_services(ws):
    cols = detect_columns(ws)
    start_row = cols["header_row"] + 1

    services = []
    for r in range(start_row, ws.max_row + 1):
        service = normalize(ws.cell(r, cols["service_col"]).value)
        price_value = ws.cell(r, cols["price_col"]).value
        sub_cat = normalize(ws.cell(r, cols["sub_col"]).value) if cols["sub_col"] else ""
        main_cat = normalize(ws.cell(r, cols["main_col"]).value) if cols["main_col"] else ""

        if not service:
            continue
        if service.lower() in {"الخدمه", "الخدمة", "service_name", "service"}:
            continue

        if price_value is None or normalize(price_value) == "":
            continue

        try:
            price = float(price_value)
        except Exception:
            continue

        services.append(
            {
                "service": service,
                "price": price,
                "sub_cat": sub_cat,
                "main_cat": main_cat
            }
        )

    return services


def write_import_template(output_path, services, classifier):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing_Template"
    ws.sheet_view.rightToLeft = True

    headers = [
        "service_name / اسم الخدمة ★",
        "service_code / الكود",
        "contract_price / سعر العقد",
        "main_category / التصنيف الرئيسي",
        "sub_category / البند (التصنيف الفرعي)",
        "notes / ملاحظات"
    ]

    for i, header in enumerate(headers, start=1):
        ws.cell(1, i, header)

    example = ["فحص شامل", "MC-001", 100.0, "عيادات خارجية", "عام", "مثال - احذف هذا الصف"]
    for i, value in enumerate(example, start=1):
        ws.cell(2, i, value)

    row = 3
    for item in services:
        service_name = item["service"]
        service_code = extract_code(service_name)
        classification = classifier.classify(service_name, item["main_cat"], item["sub_cat"])
        system_main = classification["main"]
        system_sub = classification["sub"]
        contract_price = item["price"]

        data = [service_name, service_code, contract_price, system_main, system_sub, ""]
        for i, value in enumerate(data, start=1):
            ws.cell(row, i, value)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40

    wb.save(output_path)


def evaluate_accuracy(services, classifier):
    total = len(services)
    if total == 0:
        return {
            "total_services": 0,
            "recognized_main_ratio": 0.0,
            "recognized_sub_ratio": 0.0,
            "both_recognized_ratio": 0.0,
            "overall_confidence_score": 0.0,
            "unknown_main_categories": [],
            "unknown_sub_categories": []
        }

    recognized_main = 0
    recognized_sub = 0
    both_recognized = 0
    unknown_main = set()
    unknown_sub = set()
    ai_used = 0
    fallback_used = 0
    confidence_sum = 0.0

    for item in services:
        result = classifier.classify(item["service"], item["main_cat"], item["sub_cat"])
        main = result["main"]
        sub = result["sub"]
        confidence_sum += float(result.get("confidence", 0.0))
        if result.get("method") == "ai":
            ai_used += 1
        else:
            fallback_used += 1

        main_ok = main in SYSTEM_MAIN_CATEGORIES
        sub_ok = sub in SYSTEM_SUB_CATEGORIES

        if main_ok:
            recognized_main += 1
        else:
            unknown_main.add(main)

        if sub_ok:
            recognized_sub += 1
        else:
            unknown_sub.add(sub)

        if main_ok and sub_ok:
            both_recognized += 1

    main_ratio = recognized_main / total
    sub_ratio = recognized_sub / total
    both_ratio = both_recognized / total

    # Taxonomy validity score + model confidence score.
    taxonomy_confidence = round((0.2 * main_ratio + 0.3 * sub_ratio + 0.5 * both_ratio) * 100, 2)
    model_confidence = round((confidence_sum / total) * 100, 2)
    confidence = min(taxonomy_confidence, model_confidence)

    return {
        "total_services": total,
        "recognized_main_ratio": round(main_ratio * 100, 2),
        "recognized_sub_ratio": round(sub_ratio * 100, 2),
        "both_recognized_ratio": round(both_ratio * 100, 2),
        "overall_confidence_score": round(confidence, 2),
        "model_avg_confidence": model_confidence,
        "ai_classified_rows": ai_used,
        "fallback_rows": fallback_used,
        "unknown_main_categories": sorted(unknown_main),
        "unknown_sub_categories": sorted(unknown_sub)
    }


def process_file(input_path, output_dir, classifier):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    services = extract_services(ws)

    accuracy = evaluate_accuracy(services, classifier)

    if not services:
        return {
            "input_file": str(input_path),
            "skipped": True,
            "reason": "no_service_rows_detected",
            "accuracy": accuracy
        }

    output_name = f"{input_path.stem}_جاهز_للاستيراد.xlsx"
    output_path = output_dir / output_name
    write_import_template(output_path, services, classifier)

    return {
        "input_file": str(input_path),
        "output_file": str(output_path),
        "accuracy": accuracy
    }


def list_candidate_files(folder):
    result = []
    for path in folder.glob("*.xlsx"):
        name = path.name.lower()
        if name.startswith("~$"):
            continue
        # Skip files that are clearly output templates/reports
        if "جاهز_للاستيراد" in name or "import_ready" in name or "نتيجة_" in name:
            continue
        # Skip known non-price files
        if "coverage" in name or "قواعد" in name or "علاقات" in name:
            continue
        # Prefer files that look like price lists
        if not ("price" in name or "قائمة" in name or "اسعار" in name or "contract" in name):
            continue
        result.append(path)
    return sorted(result)


def main():
    parser = argparse.ArgumentParser(description="Classify health facility price lists and generate import-ready templates")
    parser.add_argument("--input", type=str, help="Single input xlsx file")
    parser.add_argument("--input-dir", type=str, default="draft", help="Input folder for batch mode")
    parser.add_argument("--output-dir", type=str, default="draft/import_ready", help="Output folder")
    parser.add_argument("--all", action="store_true", help="Process all xlsx files in input-dir")
    parser.add_argument("--report", type=str, default="draft/import_ready/classification_accuracy_report.json", help="Output report JSON path")
    parser.add_argument("--ai-model", type=str, default=os.getenv("AI_MODEL", "gpt-4o-mini"), help="AI model name")
    parser.add_argument(
        "--ai-endpoint",
        type=str,
        default=os.getenv("AI_ENDPOINT", "https://api.openai.com/v1/chat/completions"),
        help="OpenAI-compatible chat completions endpoint"
    )
    parser.add_argument("--ai-api-key", type=str, default=os.getenv("OPENAI_API_KEY", ""), help="AI API key")
    parser.add_argument("--ai-timeout", type=int, default=45, help="AI request timeout in seconds")

    args = parser.parse_args()

    classifier = AIClassifier(
        model=args.ai_model,
        endpoint=args.ai_endpoint,
        api_key=args.ai_api_key,
        timeout=args.ai_timeout
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    report = {"files": [], "summary": {}}

    if args.input:
        input_file = Path(args.input)
        result = process_file(input_file, output_dir, classifier)
        report["files"].append(result)
    elif args.all:
        input_dir = Path(args.input_dir)
        files = list_candidate_files(input_dir)
        for file in files:
            try:
                result = process_file(file, output_dir, classifier)
                report["files"].append(result)
            except Exception as exc:
                report["files"].append(
                    {
                        "input_file": str(file),
                        "error": str(exc)
                    }
                )
    else:
        raise SystemExit("Specify --input <file> or --all")

    # Summary metrics
    processed = [f for f in report["files"] if "accuracy" in f and f["accuracy"]["total_services"] > 0]
    skipped = [f for f in report["files"] if f.get("skipped")]
    if processed:
        avg_confidence = sum(f["accuracy"]["overall_confidence_score"] for f in processed) / len(processed)
        total_services = sum(f["accuracy"]["total_services"] for f in processed)
        report["summary"] = {
            "processed_files": len(processed),
            "skipped_files": len(skipped),
            "total_services": total_services,
            "average_confidence_score": round(avg_confidence, 2)
        }
    else:
        report["summary"] = {
            "processed_files": 0,
            "skipped_files": len(skipped),
            "total_services": 0,
            "average_confidence_score": 0.0
        }

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], ensure_ascii=False))
    print(json.dumps({
        "ai_enabled": classifier.enabled,
        "ai_model": classifier.model,
        "ai_endpoint": classifier.endpoint
    }, ensure_ascii=False))
    print(f"Report saved to: {report_path}")


if __name__ == "__main__":
    main()
