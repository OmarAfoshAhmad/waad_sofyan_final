import openpyxl
p = r'draft/اسعار مستشفى الصفوة.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb[wb.sheetnames[0]]
print('sheet', ws.title, 'rows', ws.max_row, 'cols', ws.max_column)
for r in range(1, 90):
    vals = [ws.cell(r, c).value for c in range(1, 13)]
    non_empty = [v for v in vals if v is not None and str(v).strip() != '']
    if non_empty:
        print(r, vals)
