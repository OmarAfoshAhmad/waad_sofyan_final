import os
import glob
from openpyxl import load_workbook

v3_folder = r'D:\tba_waad_system-main_success\tba_waad_system-main\Unified_Benefit_Tables_Ready_V3'
excel_files = glob.glob(os.path.join(v3_folder, '*.xlsx'))

for file in excel_files:
    print(f"Processing {os.path.basename(file)}...")
    try:
        wb = load_workbook(file)
        changed = False
        for sheet in wb.sheetnames:
            if sheet == 'قواعد التغطية':
                wb[sheet].title = 'قواعد_التغطية'
                changed = True
        
        if changed:
            wb.save(file)
            print(f"  -> Fixed sheet name in {os.path.basename(file)}")
        else:
            print(f"  -> No changes needed.")
    except Exception as e:
        print(f"  -> Error processing {os.path.basename(file)}: {e}")

print("Done fixing all V3 files.")
