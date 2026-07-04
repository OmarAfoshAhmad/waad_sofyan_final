import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── التصنيفات المعتمدة ───────────────────────────────────────────
CATEGORIES = {
    "CAT001": ("الايواء و العلاج غرفة خاصة", "ايواء"),
    "CAT002": ("الدواء و المستلزمات الطبية", "ايواء"),
    "CAT003": ("العناية الفائقة و عناية القلب", "ايواء"),
    "CAT004": ("رسوم الاطباء و الجراحين و المستشارين", "ايواء"),
    "CAT005": ("الكشوف التشخيصية", "ايواء"),
    "CAT006": ("علاج الاسنان بالطوارئ - داخل المستشفى", "ايواء"),
    "CAT007": ("الاسعاف المحلي", "ايواء"),
    "CAT008": ("التمريض في المنزل او النقاهة", "ايواء"),
    "CAT009": ("العلاج طبيعي", "ايواء"),
    "CAT010": ("تكلفة اصابات العمل", "ايواء"),
    "CAT011": ("التصوير بالرنين المغناطيسي و المقطعي و الطبقي", "ايواء"),
    "CAT012": ("التصوير بالاشعة و تحليل العينات و الفحوص التشخيصية", "ايواء"),
    "CAT013": ("زرع الاعضاء", "ايواء"),
    "CAT014": ("الطب النفسي", "ايواء"),
    "CAT015": ("جراحة للمريض خارج المستشفى", "ايواء"),
    "CAT016": ("الاورام", "ايواء"),
    "CAT017": ("الغسيل الكلوي", "ايواء"),
    "CAT018": ("الاخلاء الطبي", "ايواء"),
    "CAT019": ("تكلفة شخص مرافق", "ايواء"),
    "CAT020": ("تكلفة السفر لاحد افراد العائلة", "ايواء"),
    "CAT021": ("الولادة الطبيعية و القيصرية", "ايواء"),
    "CAT022": ("مضاعفات الحمل و الولادة", "ايواء"),
    "CAT023": ("رسوم الاخصائيين و تحاليل و مختبرات و اشعة", "عيادات خارجية"),
    "CAT024": ("العلاجات والادوية الروتينية", "عيادات خارجية"),
    "CAT025": ("الاجهزة والمعدات الطبية", "عيادات خارجية"),
    "CAT026": ("العلاج الطبيعي المقرر", "عيادات خارجية"),
    "CAT027": ("علاج الاسنان الروتيني", "عيادات خارجية"),
    "CAT028": ("كشوف العيون والبصريات", "عيادات خارجية"),
    "CAT029": ("ادوية الامراض المزمنة", "عيادات خارجية"),
}

# ─── قواعد التصنيف بالكلمات المفتاحية ───────────────────────────
RULES = [
    # ICU / عناية فائقة
    (r"عناية فائقة|icu|intensive|cardiac care|ccu|nicu|picu", "CAT003"),
    # رنين / مقطعي
    (r"رنين|mri|magnetic resonance|مقطعي|ct scan|tomography", "CAT011"),
    # غسيل كلوي
    (r"غسيل كلوي|dialysis|hemodialysis", "CAT017"),
    # اورام
    (r"ورم|اورام|سرطان|كيماوي|oncology|chemotherapy|radiotherapy|tumor|cancer", "CAT016"),
    # ولادة
    (r"ولادة|قيصرية|delivery|caesarean|cesarean|birth|نفاس", "CAT021"),
    # مضاعفات حمل
    (r"مضاعفات.*حمل|prenatal|antenatal|pregnancy", "CAT022"),
    # طب نفسي
    (r"نفسي|نفسية|psychiatr|psycholog|mental health", "CAT014"),
    # زرع اعضاء
    (r"زرع|transplant", "CAT013"),
    # اسعاف
    (r"اسعاف|ambulance", "CAT007"),
    # اخلاء
    (r"اخلاء|evacuation|medevac", "CAT018"),
    # تمريض منزل
    (r"تمريض منزل|نقاهة|home nursing|convalescent", "CAT008"),
    # اصابات عمل
    (r"اصابة عمل|occupational", "CAT010"),
    # جراحة خارجية
    (r"عملية خارج|day surgery|outpatient surgery", "CAT015"),
    # تخدير
    (r"تخدير|anesthesia", "CAT004"),
    # رسوم اطباء وجراحين - موسّع
    (r"مرور طبيب|مرور جراح|مرور اخصائي|رسوم طبيب|رسوم جراح|surgeon|doctor fee|physician|consultant fee|مناوب", "CAT004"),
    (r"^طبيب$|^اخصائي$|^استشاري$|^مرور$", "CAT004"),
    (r"زيارة طارئ|طوارئ.*طبيب|رسوم طوارئ", "CAT004"),
    # اقامة / غرف
    (r"اقامة|غرفة|سرير|vip|suite|accommodation", "CAT001"),
    # اسنان
    (r"سن|اسنان|ضرس|لثه|لثة|تلبيس|تقويم|حشو|خلع|dental|tooth|teeth|orthodont|crown|filling|extraction|scaling|root canal|فينير|تبييض|بورسلين|طاقم|سراميك.*سن", "CAT027"),
    # عيون وبصريات
    (r"عيون|عين|بصر|بصريات|نظارة|عدسة|قرنية|شبكية|ophthalm|eye|vision|optical|glaucom|cataract|cornea|retina|lasik|glasses|spectacle", "CAT028"),
    # علاج طبيعي
    (r"علاج طبيعي|physiotherap|physical therapy|جلسة علاج|تاهيل|rehabilitation|ultrasound therapy|tens|paraffin|traction", "CAT009"),
    (r"تمرينات|تمرين|تدليك|علاج يدوي|كمدات|كمادات|شمع|موجات.*سمعية|موجات.*بنفسجية|موجات.*قصيرة|تداخل كهربائي|موجات تصادمية|شد.*عنق|شد.*ظهر|جذب", "CAT009"),
    # صور سونار ودوبلر وايكو (تشخيصية - CAT012)
    (r"صورة|سونار|دوبلر|ايكو|echo|ultrasound|doppler|بطن|حوض|ثدي|خصية|رقبة|غدة درقية|مفصل|صورتين", "CAT012"),
    (r"تخطيط|ecg|ekg|eeg|audiometr|نبض الجنين|رسم قلب", "CAT012"),
    (r"اشعة|x.ray|xray|fluoroscop|mammograph|صورة.*شعة", "CAT012"),
    # تحاليل ومختبرات
    (r"تحليل|فحص دم|فحص بول|مختبر|اشعة سينية|endoscopy|منظار|laborator|blood test|urine test|culture|biopsy|radiolog|patholog", "CAT012"),
    (r"CBC|ESR|CRP|HBA1C|TSH|PSA|PCR|antigen|antibod|stool|occult|RETIC|reticulocyt|هرمون|hormone|فيرس|مسحة", "CAT012"),
    (r"صبغة|وظائف الرئة|قياس السكر|سكر.*دم", "CAT012"),
    # كشف / زيارة / استشارة خارجية / متابعة إيواء
    (r"كشف|استشارة|outpatient|consultation|clinic visit|زيارة الطبيب|متابعة|اشراف طبي", "CAT005"),
    # اجراءات جراحية وعلاجية - CAT004 (رسوم جراحين/عمليات)
    (r"استئصال|بتر|ربط|فتح|فتحة|توسيع|تفويه|كحت|ترقيع|إصلاح|اصلاح|تثبيت", "CAT004"),
    (r"ناسور|لوزتين|لحمية|زائدة|مرارة|بواسير|غدة|رحم|طحال|معدة|قولون|كبد|كلية", "CAT004"),
    (r"جبس|ديناكاست|رباط|plaster|cast|خياطة|suture|غرز|فك.*غرز|تنظيف جرح|wound|غرف عمليات|عملية|جراح|operation|procedure", "CAT004"),
    (r"كي|cauteriz|ازالة ثاليل|wart|جسم غريب|foreign body|سحب سائل|aspiration|اكياس دهنية|lipoma|عقدة|خراج", "CAT004"),
    (r"قص لسان|ختان|circumcision|تركيب لولب|iud|فك لولب|تركيب.*مهبلي", "CAT004"),
    (r"منظار|laparoscop|colonoscop|gastroscop|hysteroscop|تنظير", "CAT004"),
    (r"قيمة اضافية.*طبيب|استدعاء", "CAT004"),
    # إجراءات طبية عامة
    (r"نقل دم|غسيل معدة|قياس ضغط|غسيل أذن|غسيل اذن|ثقب الاذن|علاج بالليزر|علاج ضوئي", "CAT005"),
    (r"تغذية", "CAT002"),
    # دواء ومستلزمات - CAT002
    (r"دواء|ادوية|مستلزمات طبية|medication|medicine|supplies|consumable|مستهلكات", "CAT002"),
    (r"حقن|injection|syringe|iv fluid|ضمادة|dressing|bandage|catheter|قسطرة|غيار|زجاجة سوائل", "CAT002"),
    (r"بخار|nebulizer|اكسجين|اوكسيجين|oxygen|ventolin|atrovent|salbutamol|aminophyllin", "CAT002"),
    (r"اسطوانة|cylinder|تنفسي.*علاج|respiratory.*therapy", "CAT002"),
    (r"حقنة|امبول|ampoule|vial|فيال|tablet|capsule|حبة|كبسولة", "CAT002"),
    # اجهزة ومعدات
    (r"جهاز|معدات|equipment|device|wheelchair|عكازات|crutch|prosthe|orthotic|دعامة", "CAT025"),
    # ادوية مزمنة
    (r"مزمن|chronic", "CAT029"),
    # صيدلية
    (r"صيدل|pharmacy|وصفة|prescription", "CAT024"),
]

# ─── إعدادات المجلدات ─────────────────────────────────────────────
FOLDER_DEFAULTS = {
    "اسنان منظم":       ("CAT027", "عيادات خارجية"),
    "بصريات منظم":      ("CAT028", "عيادات خارجية"),
    "علاج طبيعي منظم": ("CAT026", "عيادات خارجية"),
    "مختبرات منظم":    ("CAT023", "عيادات خارجية"),
    "مصحات منظمة":     (None, None),
}
OP_FOLDERS = {"اسنان منظم", "بصريات منظم", "علاج طبيعي منظم", "مختبرات منظم"}

HEADER = [
    "service_name / اسم الخدمة",
    "service_code / الكود",
    "contract_price / سعر العقد",
    "main_category / التصنيف الرئيسي",
    "sub_category / البند التصنيف الفرعي",
    "notes / ملاحظات",
]


def classify_service(name_ar, name_en, folder_name):
    text = " ".join(filter(None, [
        str(name_ar) if name_ar else "",
        str(name_en) if name_en else ""
    ])).lower().strip()
    if not text or text in ("none", "nan"):
        return None, None, False
    for pattern, cat_code in RULES:
        if re.search(pattern, text, re.IGNORECASE | re.UNICODE):
            if cat_code == "CAT009" and folder_name in OP_FOLDERS:
                cat_code = "CAT026"
            name, parent = CATEGORIES[cat_code]
            return parent, cat_code + " - " + name, True
    dc, dp = FOLDER_DEFAULTS.get(folder_name, (None, None))
    if dc:
        name, parent = CATEGORIES[dc]
        return parent, dc + " - " + name, True
    return None, None, False


def make_thin():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def process_file(inp, outp, folder):
    try:
        wb = openpyxl.load_workbook(inp)
    except Exception as e:
        print("  ERROR:", e)
        return 0, 0
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0
    hdr = [str(c).strip() if c else "" for c in rows[0]]

    def fc(keys):
        for k in keys:
            for i, h in enumerate(hdr):
                if k in h:
                    return i
        return None

    col_ar = fc(["عربي", "اسم الخدمة"])
    col_en = fc(["انجليزي", "English", "انكليزي"])
    col_code = fc(["الكود", "كود", "code"])
    col_price = fc(["السعر", "سعر", "price"])
    if col_ar is None and col_en is None:
        col_ar = fc(["اسم", "name", "الخدمة"])

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "قائمة الاسعار"
    ws2.sheet_view.rightToLeft = True

    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    for ci, t in enumerate(HEADER, 1):
        c = ws2.cell(row=1, column=ci, value=t)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_thin()
    ws2.row_dimensions[1].height = 30

    total = 0
    unclass_count = 0
    for rd in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in rd):
            continue
        nar = rd[col_ar] if col_ar is not None else None
        nen = rd[col_en] if col_en is not None else None
        cod = rd[col_code] if col_code is not None else None
        pri = rd[col_price] if col_price is not None else None
        parts = []
        if nar and str(nar).strip() not in ("", "None", "nan"):
            parts.append(str(nar).strip())
        if nen and str(nen).strip() not in ("", "None", "nan"):
            parts.append(str(nen).strip())
        sname = " / ".join(parts) if parts else ""
        if not sname:
            continue
        mc, sc, ok = classify_service(nar, nen, folder)
        note = "" if ok else "غير مصنف - يحتاج مراجعة يدوية"
        if not ok:
            unclass_count += 1
        ri = ws2.max_row + 1
        for ci, val in enumerate([sname, cod, pri, mc if ok else "---", sc if ok else "---", note], 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            fc_color = "FFCCCC" if not ok else ("EBF3FB" if ri % 2 == 0 else "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fc_color)
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cell.border = make_thin()
        total += 1

    for col, width in {1: 42, 2: 15, 3: 15, 4: 20, 5: 58, 6: 38}.items():
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"
    os.makedirs(os.path.dirname(outp), exist_ok=True)
    try:
        wb2.save(outp)
    except PermissionError:
        print("    *** SKIP: الملف مفتوح في Excel - اغلق الملف وأعد التشغيل:", os.path.basename(outp))
        return -1, -1
    return total, unclass_count


BASE = r"d:\tba_waad_system-main_success\tba_waad_system-main\تحتاج تصنيف"
OUT = r"d:\tba_waad_system-main_success\tba_waad_system-main\مصنف"
FOLDERS = ["اسنان منظم", "بصريات منظم", "علاج طبيعي منظم", "مختبرات منظم", "مصحات منظمة"]

print("=" * 55)
print("نظام التصنيف التلقائي لقوائم الاسعار الطبية")
print("=" * 55)
gt = 0
gu = 0
fc2 = 0
for folder in FOLDERS:
    fp = os.path.join(BASE, folder)
    if not os.path.isdir(fp):
        print("المجلد غير موجود:", folder)
        continue
    print("\nمجلد:", folder)
    files = [f for f in os.listdir(fp) if f.endswith(".xlsx")]
    for fname in sorted(files):
        ip = os.path.join(fp, fname)
        on = fname.replace("_منظم", "").replace("_منظمة", "")
        on = os.path.splitext(on)[0] + "_مصنف.xlsx"
        op = os.path.join(OUT, folder, on)
        t, u = process_file(ip, op, folder)
        if t == -1:
            continue
        gt += t
        gu += u
        fc2 += 1
        ratio = str(round(u * 100.0 / t, 1)) if t > 0 else "0"
        print("  " + fname + ": " + str(t) + " خدمة - " + str(u) + " غير مصنف (" + ratio + ")")
print("\nالملخص:", fc2, "ملف -", gt, "خدمة -", gu, "غير مصنف")
print("الملفات في:", OUT)
