from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "قائمة التصنيفات المعتمدة النهائي_كامل.xlsx"
OUTPUT = ROOT / "مسودة_اعتماد_تصنيفات_وثائق_التغطية.xlsx"


def read_source():
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    sheet = workbook.active
    records = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        code, name, parent, active = row[:4]
        if code and name:
            records.append(
                {
                    "code": str(code).strip(),
                    "name": str(name).strip(),
                    "parent": str(parent).strip() if parent else "",
                    "active": str(active).strip() if active else "",
                }
            )
    return records


def build_proposals(source):
    by_code = {row["code"]: row for row in source}
    proposals = []

    def add(source_codes, action, proposed_name, alternative="", note=""):
        codes = source_codes.split(" + ")
        missing = [code for code in codes if code not in by_code]
        if missing:
            raise ValueError(f"Missing source category codes: {missing}")
        source_names = " + ".join(by_code[code]["name"] for code in codes)
        source_parents = " + ".join(dict.fromkeys(by_code[code]["parent"] or "بدون أب" for code in codes))
        proposals.append(
            {
                "source_codes": source_codes,
                "source_names": source_names,
                "source_parents": source_parents,
                "action": action,
                "proposed_name": proposed_name,
                "alternative": alternative,
                "final_name": proposed_name,
                "decision": "قيد المراجعة",
                "notes": note,
            }
        )

    # Root/context rows are shown in the reference sheet, not repeated as medical
    # benefit categories. Every name below is copied from, or split using only
    # words already present in, the approved source workbook.
    add("CAT-OPT", "إبقاء", by_code["CAT-OPT"]["name"])
    add("CAT-Go", "إبقاء", by_code["CAT-Go"]["name"])

    add("CAT001", "فك مركب", "الايواء")
    add("CAT001", "فك مركب", "العمليات")
    add("CAT002", "فك مركب", "الدواء")
    add("CAT002", "فك مركب", "المستلزمات الطبية")
    add("CAT003", "فك مركب", "العناية الفائقة")
    add("CAT003", "فك مركب", "عناية القلب")
    add("CAT004", "فك مركب", "رسوم الاطباء")
    add("CAT004", "فك مركب", "رسوم الجراحين")
    add("CAT004", "فك مركب", "المستشارين")
    add("CAT004", "فك مركب", "اجور الممارسين")
    add("CAT004", "فك مركب", "تفقات التخدير")
    add("CAT004", "فك مركب", "المعدات الجراحية")
    add("CAT004", "فك مركب", "المواد الجراحية")
    add("CAT005", "فك مركب", "الكشوف التشخيصية ( داخل المشفى )")
    add("CAT005", "فك مركب", "العلاج")
    add("CAT005", "فك مركب", "الرعاية اليومية")
    add("CAT006", "إبقاء", by_code["CAT006"]["name"])
    add("CAT007", "إبقاء", by_code["CAT007"]["name"])
    add("CAT008", "إبقاء", by_code["CAT008"]["name"])
    add(
        "CAT009 + CAT027",
        "توحيد مرشح",
        by_code["CAT009"]["name"],
        by_code["CAT027"]["name"],
        "اختر إحدى الصيغتين الموجودتين في الملف أو عدّل الاسم النهائي.",
    )
    add("CAT010", "إبقاء", by_code["CAT010"]["name"])
    add(
        "CAT011 + CAT024",
        "توحيد مرشح",
        by_code["CAT011"]["name"],
        by_code["CAT024"]["name"],
        "اختر إحدى الصيغتين الموجودتين في الملف أو عدّل الاسم النهائي.",
    )
    add("CAT012", "فك مركب", "التصوير بالاشعة")
    add("CAT012", "فك مركب", "تحليل العينات")
    add("CAT012", "فك مركب", "الفحوص التشخيصية")
    add("CAT013", "إبقاء", by_code["CAT013"]["name"])
    add("CAT014", "فك مركب", "الطب النفسي ( أدوية )")
    add("CAT014", "فك مركب", "الطب النفسي ( جلسات )")
    add("CAT015", "إبقاء", by_code["CAT015"]["name"])
    add("CAT016", "إبقاء", by_code["CAT016"]["name"])
    add("CAT017", "إبقاء", by_code["CAT017"]["name"])
    add("CAT018", "إبقاء", by_code["CAT018"]["name"])
    add("CAT019", "إبقاء", by_code["CAT019"]["name"])
    add("CAT020", "إبقاء", by_code["CAT020"]["name"])
    add("CAT021", "فك مركب", "الولادة الطبيعية")
    add("CAT021", "فك مركب", "الولادة القيصرية")
    add("CAT022", "إبقاء", by_code["CAT022"]["name"])
    add("CAT023", "فك مركب", "رسوم اخصائيين")
    add("CAT023", "فك مركب", "ممارسي مهنة الطب")
    add("CAT023", "فك مركب", "العلاج النفسي")
    add("CAT023", "فك مركب", "تحاليل")
    add("CAT023", "فك مركب", "مختبرات")
    add("CAT023", "فك مركب", "اشعة سينية")
    add("CAT023", "فك مركب", "اشعة تشخيصية")
    add("CAT025", "إبقاء", by_code["CAT025"]["name"])
    add("CAT026", "إبقاء", by_code["CAT026"]["name"])
    add("CAT028", "إبقاء", by_code["CAT028"]["name"])
    add("CAT029", "فك مركب", "علاج الاسنان ( تركيب )")
    add("CAT029", "فك مركب", "علاج الاسنان ( تقويم )")
    add("CAT029", "فك مركب", "علاج الاسنان ( زراعة )")
    add("CAT030", "إبقاء", by_code["CAT030"]["name"])
    return proposals


def configure_sheet(sheet, widths):
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 34
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B8C2CC")
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def generate():
    source = read_source()
    proposals = build_proposals(source)
    workbook = Workbook()
    workbook.remove(workbook.active)

    instructions = workbook.create_sheet("تعليمات الاعتماد")
    instructions.sheet_view.rightToLeft = True
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105
    rows = [
        ("الغرض", "هذه مسودة لاعتماد أسماء تصنيفات جداول وثائق التغطية قبل تنفيذها في المنظومة."),
        ("مصدر الأسماء", "قائمة التصنيفات المعتمدة النهائي_كامل.xlsx فقط."),
        ("المطلوب منك", "راجع عمود «الاسم النهائي المعتمد»، عدّله عند الحاجة، ثم اختر قرار الاعتماد لكل صف."),
        ("قيد المراجعة", "لا يُنفذ الصف في قاعدة البيانات."),
        ("معتمد", "الصيغة النهائية جاهزة للتنفيذ."),
        ("تعديل مطلوب", "اكتب الصيغة المطلوبة في الاسم النهائي والملاحظة."),
        ("مرفوض", "لن يدخل البند في التصنيفات النهائية."),
        ("ملاحظة", "CAT-IP وCAT-OP ظاهران في ورقة المرجع كسياقين، وليس كبندين طبيين في مسودة الاعتماد."),
        ("تنبيه", "لم تُرحّل أي بيانات ولم يُعدّل مخطط قاعدة البيانات أو الكود بناءً على هذه المسودة."),
    ]
    instructions.append(["البند", "البيان"])
    for row in rows:
        instructions.append(row)
    configure_sheet(instructions, {"A": 24, "B": 105})
    instructions.freeze_panes = "A2"
    instructions.auto_filter.ref = "A1:B10"

    draft = workbook.create_sheet("مسودة الاعتماد")
    headers = [
        "م",
        "الرمز/الرموز في الملف",
        "الاسم الأصلي كما في الملف",
        "التصنيف الأب في الملف",
        "الإجراء المقترح",
        "الصيغة المقترحة من ألفاظ الملف",
        "صيغة بديلة موجودة في الملف",
        "الاسم النهائي المعتمد (قابل للتعديل)",
        "قرار الاعتماد",
        "ملاحظاتك",
    ]
    draft.append(headers)
    for index, proposal in enumerate(proposals, start=1):
        draft.append(
            [
                index,
                proposal["source_codes"],
                proposal["source_names"],
                proposal["source_parents"],
                proposal["action"],
                proposal["proposed_name"],
                proposal["alternative"],
                proposal["final_name"],
                proposal["decision"],
                proposal["notes"],
            ]
        )
    configure_sheet(
        draft,
        {"A": 6, "B": 20, "C": 54, "D": 18, "E": 18, "F": 38, "G": 38, "H": 42, "I": 18, "J": 48},
    )
    for row in range(2, draft.max_row + 1):
        draft.row_dimensions[row].height = 45
        draft.cell(row, 8).fill = PatternFill("solid", fgColor="FFF2CC")
        draft.cell(row, 9).fill = PatternFill("solid", fgColor="FFF2CC")
        draft.cell(row, 10).fill = PatternFill("solid", fgColor="FFF2CC")
        draft.cell(row, 8).protection = Protection(locked=False)
        draft.cell(row, 9).protection = Protection(locked=False)
        draft.cell(row, 10).protection = Protection(locked=False)

    decisions = DataValidation(
        type="list",
        formula1='"قيد المراجعة,معتمد,تعديل مطلوب,مرفوض"',
        allow_blank=False,
    )
    decisions.error = "اختر قرارًا من القائمة."
    decisions.errorTitle = "قيمة غير معتمدة"
    draft.add_data_validation(decisions)
    decisions.add(f"I2:I{draft.max_row}")
    draft.conditional_formatting.add(
        f"A2:J{draft.max_row}",
        FormulaRule(formula=["$I2=\"معتمد\""], fill=PatternFill("solid", fgColor="E2F0D9")),
    )
    draft.conditional_formatting.add(
        f"A2:J{draft.max_row}",
        FormulaRule(formula=["$I2=\"مرفوض\""], fill=PatternFill("solid", fgColor="FCE4D6")),
    )

    original = workbook.create_sheet("مرجع الملف الأصلي")
    original.append(["الرمز", "الاسم", "التصنيف الأب", "الحالة"])
    for row in source:
        original.append([row["code"], row["name"], row["parent"], row["active"]])
    configure_sheet(original, {"A": 16, "B": 78, "C": 18, "D": 12})
    for row in range(2, original.max_row + 1):
        original.row_dimensions[row].height = 32

    review = workbook.create_sheet("ملخص المراجعة")
    review.sheet_view.rightToLeft = True
    review.append(["المؤشر", "القيمة"])
    review.append(["إجمالي البنود المقترحة", f"=COUNTA('مسودة الاعتماد'!A2:A{draft.max_row})"])
    review.append(["قيد المراجعة", f'=COUNTIF(\'مسودة الاعتماد\'!I2:I{draft.max_row},"قيد المراجعة")'])
    review.append(["معتمد", f'=COUNTIF(\'مسودة الاعتماد\'!I2:I{draft.max_row},"معتمد")'])
    review.append(["تعديل مطلوب", f'=COUNTIF(\'مسودة الاعتماد\'!I2:I{draft.max_row},"تعديل مطلوب")'])
    review.append(["مرفوض", f'=COUNTIF(\'مسودة الاعتماد\'!I2:I{draft.max_row},"مرفوض")'])
    review.append(["جاهز للتنفيذ؟", '=IF(B3=0,"نعم","لا - توجد بنود قيد المراجعة")'])
    configure_sheet(review, {"A": 34, "B": 35})

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(OUTPUT)
    return OUTPUT, len(source), len(proposals)


if __name__ == "__main__":
    output, source_count, proposal_count = generate()
    print(f"Created: {output}")
    print(f"Source categories: {source_count}")
    print(f"Approval rows: {proposal_count}")
