"""Build reviewable/importable benefit-policy workbooks from the normalized V3 files.

The approved category labels below are copied verbatim from migration V84. Legacy
compound rows become one shared bucket linked to every approved atomic category.
Uncertain rows are never guessed: they are retained in SourceTrace and Review.
"""
from pathlib import Path
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Unified_Benefit_Tables_Ready_V3"
OUTPUT = ROOT / "Prepared_Benefit_Imports_V1"

CATEGORIES = {
    "CAT-OPT": "نظارة طبية", "CAT-ROOM": "الإيواء غرفة خاصة", "CAT-SURGERY": "العمليات الجراحية",
    "CAT-DRUG": "الدواء", "CAT-MED-SUP": "المستلزمات الطبية", "CAT-ICU": "العناية الفائقة",
    "CAT-CCU": "عناية القلب", "CAT-PRACT-FEE": "رسوم الأطباء والجراحين والمستشارين والممارسين",
    "CAT-ANESTHESIA": "نفقات التخدير", "CAT-SURG-MAT": "المعدات والمواد الجراحية",
    "CAT-DIAGNOSTIC": "الكشوف التشخيصية", "CAT-DAY-CARE": "العلاج والرعاية اليومية",
    "CAT-DENT-EMERG": "علاج الاسنان بالطوارئ للمريض داخل مستشفى", "CAT-AMBULANCE": "الاسعاف المحلي",
    "CAT-HOME-NURSING": "التمريض في المنزل أو النقاهة ( بديل الاقامة بعد الخروج )",
    "CAT-PHYSIO": "العلاج الطبيعي", "CAT-IMG-ADV": "التصوير بالرنين المغناطيسي و المقطعي و الطبقي",
    "CAT-LAB": "تحاليل و مختبرات", "CAT-IMG-DIAG": "اشعة سينية و اشعة تشخيصية",
    "CAT-TRANSPLANT": "زرع الاعضاء", "CAT-PSYCH-DRUG": "الطب النفسي ( أدوية )",
    "CAT-PSYCH-SESS": "الطب النفسي ( جلسات )", "CAT-ONCOLOGY": "علاج الاورام",
    "CAT-DIALYSIS": "الغسيل الكلوي", "CAT-MAT-NORMAL": "الولادة الطبيعية",
    "CAT-MAT-CS": "الولادة القيصرية", "CAT-MAT-COMP": "مضاعفات الحمل و الولادة",
    "CAT-DME": "الاجهزه و المعدات الطبية و فق تقرير الطبيب المختص",
    "CAT-DENT-ROUTINE": "علاج الاسنان الروتيني ( كشف- خلع- حشو- تنظيف )",
    "CAT-DENT-PROSTHO": "علاج الاسنان ( تركيب )", "CAT-DENT-ORTHO": "علاج الاسنان ( تقويم )",
    "CAT-DENT-IMPLANT": "علاج الاسنان ( زراعة )", "CAT-EYE-EXAM": "كشوف العيون",
    "CAT-CARDIAC-SURGERY": "عمليات القلب والشرايين",
    "CAT-DRUG-GENERAL": "أدوية الصرف العام",
    "CAT-DRUG-CHRONIC": "أدوية الأمراض المزمنة",
    "CAT-SPEECH-THERAPY": "جلسات علاج النطق للأطفال حتى سن 16 عام",
    "CAT-FERTILITY-DRUG": "أدوية أمراض الخصوبة والعقم",
    "CAT-THERAPEUTIC-INJ": "الحقن العلاجية",
    "CAT-STIMULANT-DRUG": "إبر منشطة وأدوية وكل ما يتعلق بها",
}

# (approved category, encounter context). The list preserves compound clauses.
MAP = {
    "CAT001": [("CAT-ROOM", "INPATIENT"), ("CAT-SURGERY", "INPATIENT")],
    "CAT002": [("CAT-DRUG", "INPATIENT"), ("CAT-MED-SUP", "INPATIENT")],
    "CAT003": [("CAT-ICU", "INPATIENT"), ("CAT-CCU", "INPATIENT")],
    "CAT004": [("CAT-PRACT-FEE", "INPATIENT"), ("CAT-ANESTHESIA", "INPATIENT"), ("CAT-SURG-MAT", "INPATIENT")],
    "CAT005": [("CAT-DIAGNOSTIC", "INPATIENT"), ("CAT-DAY-CARE", "INPATIENT")],
    "CAT006": [("CAT-DENT-EMERG", "INPATIENT")], "CAT007": [("CAT-AMBULANCE", "EMERGENCY")],
    "CAT008": [("CAT-HOME-NURSING", "SPECIAL")], "CAT009": [("CAT-PHYSIO", "INPATIENT")],
    "CAT011": [("CAT-IMG-ADV", "INPATIENT")],
    "CAT012": [("CAT-LAB", "INPATIENT"), ("CAT-IMG-DIAG", "INPATIENT"), ("CAT-DIAGNOSTIC", "INPATIENT")],
    "CAT013": [("CAT-TRANSPLANT", "INPATIENT")],
    "CAT014": [("CAT-PSYCH-DRUG", "INPATIENT"), ("CAT-PSYCH-SESS", "INPATIENT")],
    "CAT015": [("CAT-SURGERY", "OUTPATIENT")],
    "CAT016": [("CAT-ONCOLOGY", "INPATIENT"), ("CAT-ONCOLOGY", "OUTPATIENT")],
    "CAT017": [("CAT-DIALYSIS", "INPATIENT"), ("CAT-DIALYSIS", "OUTPATIENT")],
    "CAT021": [("CAT-MAT-NORMAL", "INPATIENT"), ("CAT-MAT-CS", "INPATIENT")],
    "CAT022": [("CAT-MAT-COMP", "INPATIENT")],
    "CAT023": [("CAT-PRACT-FEE", "OUTPATIENT"), ("CAT-PSYCH-SESS", "OUTPATIENT"),
               ("CAT-LAB", "OUTPATIENT"), ("CAT-IMG-DIAG", "OUTPATIENT")],
    "CAT024": [("CAT-IMG-ADV", "OUTPATIENT")], "CAT025": [("CAT-DRUG", "OUTPATIENT")],
    "CAT026": [("CAT-DME", "OUTPATIENT")], "CAT027": [("CAT-PHYSIO", "OUTPATIENT")],
    "CAT028": [("CAT-DENT-ROUTINE", "OUTPATIENT")],
    "CAT029": [("CAT-DENT-PROSTHO", "OUTPATIENT"), ("CAT-DENT-ORTHO", "OUTPATIENT"), ("CAT-DENT-IMPLANT", "OUTPATIENT")],
    "CAT030": [("CAT-EYE-EXAM", "OUTPATIENT"), ("CAT-OPT", "OUTPATIENT")],
    "CAT-OPT": [("CAT-EYE-EXAM", "OUTPATIENT"), ("CAT-OPT", "OUTPATIENT")],
}
SPECIAL = {"CAT010": "BEN-WORK-INJURY", "CAT018": "BEN-EVACUATION", "CAT019": "BEN-COMPANION", "CAT020": "BEN-FAMILY-TRAVEL"}
ANNUAL_LIMITS = {"اركاديا": 50000, "المنطقة الحرة جليانة": 60000, "الواحة": 50000,
                 "توسيالي أجانب": 50000, "حجر الماس": 50000, "رواق": 50000,
                 "مصلحة الجمارك": 60000, "وثيقة اوزون": 30000, "وعد المعماري": 50000}

# Explicit decisions supplied from the source documents / policy owner.
POLICY_COVERAGE_OVERRIDES = {("الواحة", "CAT024"): 100}
POLICY_NOT_COVERED = {("الواحة", "CAT029")}

HEADERS = {
    "Rules": ["category_code", "category_name", "encounter_type", "coverage_percent", "copay_percentage", "waiting_days", "requires_preapproval", "priority", "notes", "active"],
    "Groups": ["group_code", "group_name", "context_type", "aggregation_mode", "active"],
    "Buckets": ["bucket_code", "bucket_name", "group_code", "context_type", "amount_limit", "times_limit", "days_limit", "period_type", "period_value", "counting_method", "consumption_basis", "parent_bucket_code", "shared", "active"],
    "Links": ["category_code", "encounter_type", "bucket_code", "consumption_order", "consumption_mode", "mandatory"],
    "SpecialBenefits": ["definition_code", "name", "coverage_percent", "copay_percentage", "amount_limit", "times_limit", "period_type", "requires_preapproval", "notes", "source_clause", "active"],
    "Decisions": ["legacy_code", "source_clause", "decision", "coverage_percent", "evidence"],
    "Review": ["severity", "legacy_code", "source_clause", "reason", "required_action"],
    "SourceTrace": ["legacy_code", "legacy_name", "legacy_parent", "coverage", "amount", "times", "waiting_days", "preapproval", "notes"],
}

def number(value):
    if value in (None, ""): return None
    text = str(value).replace(",", "").replace("%", "").strip()
    try: return float(text) if "." in text else int(text)
    except ValueError: return None

def period(notes):
    text = str(notes or "")
    match = re.search(r"كل\s*([235])\s*سن", text)
    if match: return "MULTI_YEAR_POLICY", int(match.group(1))
    if "شهري" in text or "شهر" in text: return "MONTHLY", 1
    return "ANNUAL", 1

def code_token(code): return re.sub(r"[^A-Z0-9]+", "-", code.upper()).strip("-")

def load_rows(path):
    ws = load_workbook(path, data_only=True).active
    rows = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        if values[0]: rows.append(tuple(values[:9]))
    return rows

def build(path):
    name, source_rows = path.stem, load_rows(path)
    wb = Workbook(); wb.remove(wb.active)
    data = {sheet: [] for sheet in HEADERS}
    review = data["Review"]
    for row in source_rows: data["SourceTrace"].append(list(row))

    annual = ANNUAL_LIMITS.get(name)
    if annual:
        data["Groups"].append(["G-GENERAL", "السقف المالي السنوي العام", "ANY", "HIERARCHICAL", True])
        data["Buckets"].append(["B-GENERAL", "السقف المالي السنوي العام", "G-GENERAL", "ANY", annual, None, None, "ANNUAL", 1, "EACH_LINE", "COMPANY_SHARE", None, True, True])
    else:
        review.append(["WARNING", None, None, "لم يُثبت السقف المالي العام من ملف V3", "راجع الوثيقة الأصلية وأضف وعاء B-GENERAL عند وجوده"])

    rule_keys, used_optical = set(), False
    for legacy, legacy_name, parent, coverage_raw, amount_raw, times_raw, waiting, preapproval, notes in source_rows:
        if legacy in ("CAT-IP", "CAT-OP", "CAT-Go"): continue
        # The unit-bank sheet is rebuilt from the scanned source image below.
        # Its V3 row mapping mixes special clauses with broad legacy categories.
        if name == "مصرف الوحدة": continue
        # CAT030 is authoritative for the combined eye clause when both duplicates exist.
        if legacy == "CAT-OPT" and any(r[0] == "CAT030" for r in source_rows): continue
        coverage, amount, times = number(coverage_raw), number(amount_raw), number(times_raw)
        source_clause = legacy_name + (f" — {notes}" if notes else "")
        decision_key = (name, legacy)
        if decision_key in POLICY_NOT_COVERED:
            data["Decisions"].append([legacy, source_clause, "NOT_COVERED", 0, "اعتماد مالك الوثيقة: غير مغطى"])
            continue
        if decision_key in POLICY_COVERAGE_OVERRIDES:
            coverage = POLICY_COVERAGE_OVERRIDES[decision_key]
            data["Decisions"].append([legacy, source_clause, "COVERED", coverage, "نص الوثيقة: Full cover"])
        if legacy in SPECIAL:
            if coverage is None:
                review.append(["BLOCKER", legacy, source_clause, "نسبة التغطية غير مذكورة", "أدخل النسبة بعد الرجوع للوثيقة"]); continue
            data["SpecialBenefits"].append([SPECIAL[legacy], legacy_name, coverage, 100-coverage, amount, times, "ANNUAL", bool(preapproval), notes, source_clause, True])
            continue
        targets = MAP.get(legacy)
        if not targets:
            review.append(["BLOCKER", legacy, source_clause, "لا يوجد تفكيك معتمد لهذا البند", "اربطه بتصنيف معتمد أو بمنفعة خاصة بعد المراجعة"]); continue
        if coverage is None:
            review.append(["BLOCKER", legacy, source_clause, "نسبة التغطية غير مذكورة؛ لم تُخمن", "أدخل النسبة ثم أعد التوليد أو أضف القواعد يدويًا"]); continue

        token = code_token(legacy)
        bucket_code = None
        has_limit = amount is not None or times is not None
        if has_limit:
            bucket_code, group_code = f"B-{token}", f"G-{token}"
            mode = "SHARED" if len(targets) > 1 else "INDIVIDUAL"
            context = targets[0][1] if len({x[1] for x in targets}) == 1 else "ANY"
            data["Groups"].append([group_code, legacy_name, context, mode, True])
            ptype, pvalue = period(notes)
            data["Buckets"].append([bucket_code, legacy_name, group_code, context, amount, times, None, ptype, pvalue,
                                    "EACH_LINE", "COMPANY_SHARE", "B-GENERAL" if annual else None, len(targets) > 1, True])
        for cat, context in targets:
            key = (cat, context)
            if key in rule_keys:
                review.append(["WARNING", legacy, source_clause, f"تكرر {cat}/{context}; احتُفظ بأول قاعدة", "راجع أي بندين يجب أن تكون لهما شروط مختلفة"]); continue
            rule_keys.add(key)
            data["Rules"].append([cat, CATEGORIES[cat], context, coverage, 100-coverage, number(waiting), bool(preapproval), 100, f"المصدر: {source_clause}", True])
            target_bucket = bucket_code
            # A combined eye row limits glasses, not the medical eye examination, unless the source says otherwise.
            if legacy in ("CAT030", "CAT-OPT") and cat == "CAT-EYE-EXAM": target_bucket = "B-GENERAL" if annual else None
            if target_bucket: data["Links"].append([cat, context, target_bucket, 1, "PRIMARY", True])
            elif annual: data["Links"].append([cat, context, "B-GENERAL", 1, "PRIMARY", True])

    if name == "مصرف الوحدة":
        def unit_bucket(code, label, context, amount=None, times=None, ptype="ANNUAL", pvalue=1, shared=False):
            data["Groups"].append([f"G-{code}", label, context, "SHARED" if shared else "INDIVIDUAL", True])
            data["Buckets"].append([f"B-{code}", label, f"G-{code}", context, amount, times, None, ptype, pvalue,
                                    "EACH_LINE", "COMPANY_SHARE", None, shared, True])
            return f"B-{code}"

        def unit_rule(cat, context, coverage, clause, amount=None, times=None, ptype="ANNUAL", pvalue=1, preapproval=True):
            key = (cat, context)
            if key in rule_keys: return
            rule_keys.add(key)
            data["Rules"].append([cat, CATEGORIES[cat], context, coverage, 100-coverage, None, preapproval, 100, f"المصدر: {clause}", True])
            if amount is not None or times is not None:
                token = f"UNIT-{code_token(cat)}-{context}"
                data["Groups"].append([f"G-{token}", clause, context, "INDIVIDUAL", True])
                data["Buckets"].append([f"B-{token}", clause, f"G-{token}", context, amount, times, None, ptype, pvalue,
                                        "EACH_LINE", "COMPANY_SHARE", None, False, True])
                data["Links"].append([cat, context, f"B-{token}", 1, "PRIMARY", True])

        def unit_linked_rule(cat, context, coverage, clause, bucket_code, preapproval=True, notes=None):
            key = (cat, context)
            if key in rule_keys: return
            rule_keys.add(key)
            final_notes = notes or f"المصدر: {clause}"
            data["Rules"].append([cat, CATEGORIES[cat], context, coverage, 100-coverage, None, preapproval, 100, final_notes, True])
            data["Links"].append([cat, context, bucket_code, 1, "PRIMARY", True])

        unit_rule("CAT-SURGERY", "OUTPATIENT", 100, "العمليات الجراحية الصغرى (عيادات خارجية)", preapproval=False)
        unit_rule("CAT-SURGERY", "INPATIENT", 100, "العمليات الجراحية وأنواعها", amount=15000)
        unit_rule("CAT-CARDIAC-SURGERY", "INPATIENT", 100, "عمليات القلب والشرايين", amount=20000)
        unit_rule("CAT-ICU", "INPATIENT", 100, "الإيواء في العناية", amount=30000)
        unit_rule("CAT-ONCOLOGY", "INPATIENT", 75, "الأورام في الجسم", amount=15000,
                  preapproval=True)
        unit_rule("CAT-PHYSIO", "INPATIENT", 100, "العلاج الطبيعي", times=70)
        unit_rule("CAT-SPEECH-THERAPY", "OUTPATIENT", 100, "جلسات علاج النطق للأطفال حتى سن 16 عام", times=70)
        unit_rule("CAT-OPT", "OUTPATIENT", 100, "نظارة كل سنتين بسقف 500 دينار", amount=500, times=1,
                  ptype="MULTI_YEAR_POLICY", pvalue=2)
        unit_rule("CAT-DENT-PROSTHO", "OUTPATIENT", 80, "تركيبات الأسنان: تغطية 80% من قيمة الفاتورة")
        unit_rule("CAT-DRUG-CHRONIC", "OUTPATIENT", 100, "أدوية الأمراض المزمنة 300 دينار شهريا", amount=300, ptype="MONTHLY", preapproval=False)
        unit_rule("CAT-FERTILITY-DRUG", "OUTPATIENT", 100, "أدوية أمراض الخصوبة والعقم 5000 كل 3 سنوات", amount=5000,
                  ptype="MULTI_YEAR_POLICY", pvalue=3)
        unit_rule("CAT-THERAPEUTIC-INJ", "OUTPATIENT", 100, "الحقن العلاجية: 3000 للمرة الواحدة لعدد 4 مرات في السنة",
                  amount=3000, times=4)
        unit_rule("CAT-STIMULANT-DRUG", "OUTPATIENT", 100, "إبر منشطة وأدوية وكل ما يتعلق بها 6000 سنويا", amount=6000)
        unit_rule("CAT-DME", "OUTPATIENT", 100, "السماعات الطبية 5000 كل 5 سنوات", amount=5000, times=1,
                  ptype="MULTI_YEAR_POLICY", pvalue=5)
        data["SpecialBenefits"].append(["BEN-COMPANION", "تكلفة شخص مرافق واحد للشخص الذي تم اخلاءه", 100, 0,
                                        None, None, "POLICY_PERIOD", True, "إذا دعت الضرورة ذلك",
                                        "خدمات المرافق إذا دعت الضرورة ذلك", True])

        general_bucket = unit_bucket("UNIT-GENERAL-PHARMACY", "كشوفات طبية وأدوية الصرف العام", "OUTPATIENT",
                                     amount=600, shared=True)
        unit_linked_rule("CAT-DIAGNOSTIC", "OUTPATIENT", 100, "كشوفات طبية وأدوية الصرف العام", general_bucket,
                         preapproval=False)
        unit_linked_rule("CAT-DRUG-GENERAL", "OUTPATIENT", 100, "كشوفات طبية وأدوية الصرف العام", general_bucket,
                         preapproval=False)

        # Exact outpatient clause from the image.
        for cat in ("CAT-PRACT-FEE", "CAT-LAB", "CAT-IMG-DIAG", "CAT-IMG-ADV"):
            unit_rule(cat, "OUTPATIENT", 100, "الأشعة والصور والمناظير والتحاليل والكشوفات: تغطية كاملة", preapproval=False)
        review.append(["WARNING", "SOURCE-IMAGE", "الأورام في الجسم 15000 مع مشاركة 25% لغير الموظف",
                       "النظام لا يميز نسبة تغطية الموظف عن التابع داخل القاعدة نفسها؛ استُخدمت نسبة 75% كشرط محافظ",
                       "إذا أُضيف شرط أهلية حسب صفة المستفيد لاحقًا فافصلها إلى قاعدتين"])

    info = wb.create_sheet("Instructions")
    instructions = [
        ["ملف استيراد بنية وثيقة التغطية", name],
        ["الحالة", "جاهز للفحص الآلي، ويجب إغلاق عناصر BLOCKER في ورقة Review قبل الاعتماد"],
        ["قاعدة التسمية", "أسماء التصنيفات من القائمة المعتمدة/V84 حرفيًا؛ أسماء البنود الأصلية محفوظة في notes وSourceTrace"],
        ["الأوعية المشتركة", "كل بند مركب له سقف واحد أصبح وعاءً مشتركًا، وليس سقفًا مكررًا لكل تصنيف"],
        ["طريقة الاستخدام", "ارفع الملف من تبويب مجموعات المنافع والسقوف، اختر فحص الملف، ثم الاعتماد عند خلو الأخطاء"],
    ]
    for row in instructions: info.append(row)
    info.column_dimensions["A"].width = 28; info.column_dimensions["B"].width = 110
    for sheet, headers in HEADERS.items():
        ws = wb.create_sheet(sheet); ws.sheet_view.rightToLeft = True; ws.append(headers)
        for row in data[sheet]: ws.append(row)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="167D6D")
        for column in ws.columns:
            width = min(65, max(12, max(len(str(c.value or "")) for c in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width
        for row in ws.iter_rows():
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    OUTPUT.mkdir(exist_ok=True)
    target = OUTPUT / f"استيراد_{name}.xlsx"; wb.save(target)
    return target, len(data["Rules"]), len(data["Buckets"]), len(review)

if __name__ == "__main__":
    results = [build(path) for path in sorted(SOURCE.glob("*.xlsx"))]
    for target, rules, buckets, reviews in results:
        print(f"{target.name}: rules={rules}, buckets={buckets}, review={reviews}")
    summary = Workbook(); ws = summary.active; ws.title = "Summary"; ws.sheet_view.rightToLeft = True
    ws.append(["الوثيقة", "ملف الاستيراد", "عدد القواعد", "عدد الأوعية", "عناصر المراجعة", "BLOCKER", "WARNING"])
    details = summary.create_sheet("ReviewDetails"); details.sheet_view.rightToLeft = True
    details.append(["الوثيقة", "severity", "legacy_code", "source_clause", "reason", "required_action"])
    for target, rules, buckets, reviews in results:
        source = load_workbook(target, read_only=True, data_only=True)["Review"]
        rows = [r for r in source.iter_rows(min_row=2, values_only=True) if r[0]]
        blockers = sum(r[0] == "BLOCKER" for r in rows); warnings = sum(r[0] == "WARNING" for r in rows)
        document = target.stem.removeprefix("استيراد_")
        ws.append([document, target.name, rules, buckets, reviews, blockers, warnings])
        for row in rows: details.append([document, *row])
    for sheet in (ws, details):
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="167D6D")
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(70, max(14, max(len(str(c.value or "")) for c in column) + 2))
        for row in sheet.iter_rows():
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.save(OUTPUT / "ملخص_مراجعة_ملفات_الاستيراد.xlsx")
