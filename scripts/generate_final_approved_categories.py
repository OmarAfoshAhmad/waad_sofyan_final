from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "التصنيفات_النهائية_المعتمدة_للوثائق.xlsx"
IMPORT_OUTPUT = ROOT / "قائمة_التصنيفات_جاهز_للاستيراد.xlsx"
IMPORT_OUTPUT_FALLBACK = ROOT / "قائمة_التصنيفات_جاهز_للاستيراد_محدث.xlsx"

SERVICE_CATEGORIES = [
    ("CAT-OPT", "نظارة طبية", "OUTPATIENT"),
    ("CAT-ROOM", "الإيواء غرفة خاصة", "INPATIENT"),
    ("CAT-SURGERY", "العمليات الجراحية", "OUTPATIENT + INPATIENT"),
    ("CAT-DRUG", "الدواء", "OUTPATIENT + INPATIENT"),
    ("CAT-MED-SUP", "المستلزمات الطبية", "OUTPATIENT + INPATIENT"),
    ("CAT-ICU", "العناية الفائقة", "INPATIENT"),
    ("CAT-CCU", "عناية القلب", "INPATIENT"),
    ("CAT-PRACT-FEE", "رسوم الأطباء والجراحين والمستشارين والممارسين", "OUTPATIENT + INPATIENT"),
    ("CAT-ANESTHESIA", "نفقات التخدير", "OUTPATIENT + INPATIENT"),
    ("CAT-SURG-MAT", "المعدات والمواد الجراحية", "OUTPATIENT + INPATIENT"),
    ("CAT-DIAGNOSTIC", "الكشوف التشخيصية", "OUTPATIENT + INPATIENT"),
    ("CAT-DAY-CARE", "العلاج والرعاية اليومية", "INPATIENT"),
    ("CAT-DENT-EMERG", "علاج الاسنان بالطوارئ للمريض داخل مستشفى", "INPATIENT"),
    ("CAT-AMBULANCE", "الاسعاف المحلي", "EMERGENCY"),
    ("CAT-HOME-NURSING", "التمريض في المنزل أو النقاهة ( بديل الاقامة بعد الخروج )", "SPECIAL"),
    ("CAT-PHYSIO", "العلاج الطبيعي", "OUTPATIENT + INPATIENT"),
    ("CAT-IMG-ADV", "التصوير بالرنين المغناطيسي و المقطعي و الطبقي", "OUTPATIENT + INPATIENT"),
    ("CAT-LAB", "تحاليل و مختبرات", "OUTPATIENT + INPATIENT"),
    ("CAT-IMG-DIAG", "اشعة سينية و اشعة تشخيصية", "OUTPATIENT + INPATIENT"),
    ("CAT-TRANSPLANT", "زرع الاعضاء", "INPATIENT"),
    ("CAT-PSYCH-DRUG", "الطب النفسي ( أدوية )", "OUTPATIENT + INPATIENT"),
    ("CAT-PSYCH-SESS", "الطب النفسي ( جلسات )", "OUTPATIENT + INPATIENT"),
    ("CAT-ONCOLOGY", "علاج الاورام", "OUTPATIENT + INPATIENT"),
    ("CAT-DIALYSIS", "الغسيل الكلوي", "OUTPATIENT + INPATIENT"),
    ("CAT-MAT-NORMAL", "الولادة الطبيعية", "INPATIENT"),
    ("CAT-MAT-CS", "الولادة القيصرية", "INPATIENT"),
    ("CAT-MAT-COMP", "مضاعفات الحمل و الولادة", "OUTPATIENT + INPATIENT"),
    ("CAT-DME", "الاجهزه و المعدات الطبية و فق تقرير الطبيب المختص", "OUTPATIENT"),
    ("CAT-DENT-ROUTINE", "علاج الاسنان الروتيني ( كشف- خلع- حشو- تنظيف )", "OUTPATIENT"),
    ("CAT-DENT-PROSTHO", "علاج الاسنان ( تركيب )", "OUTPATIENT"),
    ("CAT-DENT-ORTHO", "علاج الاسنان ( تقويم )", "OUTPATIENT"),
    ("CAT-DENT-IMPLANT", "علاج الاسنان ( زراعة )", "OUTPATIENT"),
    ("CAT-EYE-EXAM", "كشوف العيون", "OUTPATIENT"),
    ("CAT-CARDIAC-SURGERY", "عمليات القلب والشرايين", "INPATIENT"),
    ("CAT-DRUG-GENERAL", "أدوية الصرف العام", "OUTPATIENT"),
    ("CAT-DRUG-CHRONIC", "أدوية الأمراض المزمنة", "OUTPATIENT"),
    ("CAT-SPEECH-THERAPY", "جلسات علاج النطق للأطفال حتى سن 16 عام", "OUTPATIENT"),
    ("CAT-FERTILITY-DRUG", "أدوية أمراض الخصوبة والعقم", "OUTPATIENT"),
    ("CAT-THERAPEUTIC-INJ", "الحقن العلاجية", "OUTPATIENT"),
    ("CAT-STIMULANT-DRUG", "إبر منشطة وأدوية وكل ما يتعلق بها", "OUTPATIENT"),
]

SPECIAL_BENEFITS = [
    ("BEN-WORK-INJURY", "تكلفة اصابات العمل"),
    ("BEN-EVACUATION", "الاخلاء الطبي"),
    ("BEN-COMPANION", "تكلفة شخص مرافق واحد للشخص الذي تم اخلاءه"),
    ("BEN-FAMILY-TRAVEL", "تكلفة السفر لاحد افراد عائلة المؤمن عليه في حالة الاخلاء"),
    ("BEN-RARE", "خدمات نادرة"),
]

LEGACY_CATEGORY_CODES = {"CAT-IP", "CAT-OP", "CAT-Go"}
LEGACY_CATEGORY_PATTERN = re.compile(r"CAT\d{3}")


def format_sheet(sheet, widths):
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    thin = Side(style="thin", color="B7C9D6")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def main():
    assert_no_legacy_categories()
    workbook = build_approved_workbook()
    workbook.save(OUTPUT)

    import_workbook = build_approved_workbook(include_reference_sheets=False)
    try:
        import_workbook.save(IMPORT_OUTPUT)
        saved_import = IMPORT_OUTPUT
    except PermissionError:
        import_workbook.save(IMPORT_OUTPUT_FALLBACK)
        saved_import = IMPORT_OUTPUT_FALLBACK

    print(OUTPUT)
    print(saved_import)


def assert_no_legacy_categories():
    legacy = [
        code for code, _name, _contexts in SERVICE_CATEGORIES
        if code in LEGACY_CATEGORY_CODES or LEGACY_CATEGORY_PATTERN.fullmatch(code)
    ]
    if legacy:
        raise ValueError("Legacy medical categories are not allowed in the approved workbook: " + ", ".join(legacy))


def build_approved_workbook(include_reference_sheets=True):
    workbook = Workbook()
    services = workbook.active
    services.title = "تصنيفات الخدمات المعتمدة"
    services.append(["م", "الكود المعتمد", "الاسم النهائي المعتمد", "السياقات المسموحة", "الحالة"])
    for index, (code, name, contexts) in enumerate(SERVICE_CATEGORIES, 1):
        services.append([index, code, name, contexts, "معتمد"])
    format_sheet(services, {"A": 7, "B": 24, "C": 72, "D": 30, "E": 14})

    if not include_reference_sheets:
        return workbook

    # Machine-readable compatibility sheet for the current system importer.
    # Row 2 is intentionally reserved because the shared importer treats it as
    # an example row and starts importing from Excel row 3.
    data = workbook.create_sheet("Data")
    data.append(["* رمز التصنيف\ncode", "* اسم التصنيف\nname", "رمز التصنيف الأب\nparent_code", "نشط\nactive"])
    data.append(["", "", "", ""])
    for code, name, _contexts in SERVICE_CATEGORIES:
        data.append([code, name, "", "نعم"])
    format_sheet(data, {"A": 26, "B": 76, "C": 25, "D": 14})
    data.sheet_state = "hidden"

    benefits = workbook.create_sheet("المنافع المالية الخاصة")
    benefits.append(["م", "الكود المعتمد", "الاسم النهائي المعتمد", "النوع", "الحالة"])
    for index, (code, name) in enumerate(SPECIAL_BENEFITS, 1):
        benefits.append([index, code, name, "منفعة مالية خاصة", "معتمد"])
    format_sheet(benefits, {"A": 7, "B": 26, "C": 78, "D": 24, "E": 14})

    notes = workbook.create_sheet("قرارات معمارية")
    notes.append(["القرار", "التفصيل"])
    decisions = [
        ("السياق", "OUTPATIENT وINPATIENT ليسا تصنيفين؛ يحددان قاعدة الوثيقة المطبقة."),
        ("التصنيف", "خدمة قائمة الأسعار ترتبط بتصنيف طبي واحد فقط."),
        ("التغطية", "يمكن إنشاء قاعدتين للتصنيف نفسه عند اختلاف السياق."),
        ("السقوف", "تجمع التصنيفات في مجموعات منافع وأوعية سقوف دون تغيير التصنيف الطبي."),
        ("المنافع الخاصة", "الإخلاء والمرافق والسفر وإصابات العمل لا تعامل كخدمات طبية في قوائم الأسعار."),
        ("الأسنان", "التركيب والتقويم والزراعة تصنيفات منفصلة لاختلاف قواعدها المالية."),
        ("الطب النفسي", "الأدوية والجلسات منفصلان ويمكن جمعهما في مجموعة منفعة واحدة."),
    ]
    for decision in decisions:
        notes.append(decision)
    format_sheet(notes, {"A": 28, "B": 110})

    return workbook


if __name__ == "__main__":
    main()
